import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# APP CONFIG
# ============================================================
APP_TITLE = "Codexid"
MAX_MASS_FILES = 50
MAX_TOTAL_UPLOAD_MB = 200
BIGSELLER_MAX_ROWS_PER_FILE = 10000
DEFAULT_TONGLE_GUDANGS = ["JKT-1A", "JKT-3B", "JKT-3C", "JKT-4B"]
STOCK_PRICELIST_SHEETS = ["LAPTOP", "TELCO", "PC HOM ELE", "SOF COM SUP", "ACC"]

st.set_page_config(page_title=APP_TITLE, layout="wide")

# Global style: blue loading bar
st.markdown("""
<style>
    div.stProgress > div > div > div > div {
        background-color: #2563eb;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# SESSION STATE
# ============================================================
SESSION_DEFAULTS = {
    "download_cache": {},
    "summary_cache": {},
    "stock_shopee_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_tiktokshop_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_bigseller_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_blibli_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_akulaku_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
}
for _k, _v in SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ============================================================
# GENERIC HELPERS
# ============================================================
def s(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def s_clean(x) -> str:
    if x is None:
        return ""
    txt = str(x).replace("\xa0", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def su(x) -> str:
    return s_clean(x).upper()


def norm_sku(v) -> str:
    txt = su(v)
    if not txt:
        return ""
    if re.fullmatch(r"\d+\.0", txt):
        txt = txt[:-2]
    txt = re.sub(r"\s+", "", txt)
    return txt


def split_sku_addons(full_sku: str) -> Tuple[str, List[str]]:
    parts = [p.strip() for p in s_clean(full_sku).split("+") if p and s_clean(p)]
    if not parts:
        return "", []
    return parts[0], parts[1:]


def normalize_addon_code(x) -> str:
    return su(x)


def parse_number_like_id(x) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if pd.isna(x):
            return ""
        if x.is_integer():
            return str(int(x))
        return str(x)
    return s_clean(x)


def to_int_or_none(v) -> Optional[int]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if pd.isna(v):
            return None
        return int(round(v))
    digits = re.findall(r"\d+", s_clean(v))
    if not digits:
        return None
    return int("".join(digits))


def parse_price_cell(val) -> Optional[int]:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and pd.isna(val):
                return None
            return int(round(float(val)))
        except Exception:
            return None

    txt = s_clean(val)
    if not txt:
        return None
    txt = txt.replace("Rp", "").replace("rp", "").replace(" ", "")

    if "." in txt and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "." in txt and "," not in txt:
        txt = txt.replace(".", "")
    elif "," in txt and "." not in txt:
        txt = txt.replace(",", "")

    try:
        return int(round(float(txt)))
    except Exception:
        return None


def apply_multiplier_if_needed(x: Optional[int], threshold: int = 1_000_000, multiplier: int = 1000) -> int:
    if x is None:
        return 0
    if x < threshold:
        return int(x) * multiplier
    return int(x)


def total_upload_size_mb(files: List[Any]) -> float:
    total = 0
    for f in files:
        try:
            total += len(f.getvalue())
        except Exception:
            pass
    return total / (1024 * 1024)


def lower_map_headers(ws: Worksheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = s_clean(v).lower()
        if key and key not in m:
            m[key] = c
    return m


def get_header_col_fuzzy(ws: Worksheet, header_row: int, candidates: List[str]) -> Optional[int]:
    m = lower_map_headers(ws, header_row)
    normalized = {re.sub(r"[^a-z0-9]", "", k): v for k, v in m.items()}
    for cand in candidates:
        target = re.sub(r"[^a-z0-9]", "", s_clean(cand).lower())
        if target in normalized:
            return normalized[target]
    return None


def find_header_row_by_exact(ws: Worksheet, header_text: str, scan_rows: int = 150) -> Optional[int]:
    target = su(header_text)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == target:
                return r
    return None


def find_row_contains(ws: Worksheet, needle: str, scan_rows: int = 300) -> Optional[int]:
    target = su(needle)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = su(ws.cell(r, c).value)
            if v and (target == v or target in v):
                return r
    return None


def get_first_sheet(wb) -> Worksheet:
    return wb[wb.sheetnames[0]]


def build_merged_lookup_map(ws: Worksheet) -> Dict[Tuple[int, int], object]:
    merged_map: Dict[Tuple[int, int], object] = {}
    for mr in ws.merged_cells.ranges:
        top_left_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = top_left_val
    return merged_map


def get_cell_or_merged_value(ws: Worksheet, merged_map: Dict[Tuple[int, int], object], row: int, col: int):
    v = ws.cell(row, col).value
    if v not in (None, ""):
        return v
    return merged_map.get((row, col))


def safe_set_cell_value(ws: Worksheet, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def zip_named_files(named_files: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in named_files:
            zf.writestr(fname, fbytes)
    return buf.getvalue()


def make_issues_workbook(issues: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "issues_report"
    headers = ["file", "row", "sku_full", "old_value", "new_value", "reason"]
    ws.append(headers)
    for item in issues:
        ws.append([
            item.get("file", ""),
            item.get("row", ""),
            item.get("sku_full", ""),
            item.get("old_value", ""),
            item.get("new_value", ""),
            item.get("reason", ""),
        ])
    return workbook_to_bytes(wb)


def render_summary(title: str, summary: Dict[str, Any]):
    st.subheader(title)
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Jumlah File", int(summary.get("files_total", 0)))
    c2.metric("Baris Diperiksa", int(summary.get("rows_scanned", 0)))
    c3.metric("Baris Diupdate", int(summary.get("rows_written", 0)))
    c4.metric("Skip / Issue", int(summary.get("rows_unmatched", 0) + summary.get("issues_count", 0)))


def cache_downloads(
    cache_key: str,
    result_name: str,
    result_bytes: Optional[bytes],
    issues_bytes: Optional[bytes],
    summary: Optional[Dict[str, Any]] = None,
    issues_name: str = "issues_report.xlsx",
):
    st.session_state.download_cache[cache_key] = {
        "result_name": result_name,
        "result_bytes": result_bytes,
        "issues_name": issues_name,
        "issues_bytes": issues_bytes,
    }
    if summary is not None:
        st.session_state.summary_cache[cache_key] = summary


def render_downloads(cache_key: str):
    payload = st.session_state.download_cache.get(cache_key)
    if not payload:
        return
    if payload.get("result_bytes"):
        st.download_button(
            "Download Hasil",
            payload["result_bytes"],
            file_name=payload["result_name"],
            key=f"dl_{cache_key}_result",
        )
    if payload.get("issues_bytes"):
        st.download_button(
            "Download Issues",
            payload["issues_bytes"],
            file_name=payload["issues_name"],
            key=f"dl_{cache_key}_issues",
        )


def render_cached_summary(cache_key: str, title: str = "Ringkasan Hasil"):
    summary = st.session_state.summary_cache.get(cache_key)
    if summary:
        render_summary(title, summary)


def get_change_sheet(wb):
    for sname in wb.sheetnames:
        if su(sname) == "CHANGE":
            return wb[sname]
    raise ValueError("Sheet 'CHANGE' tidak ditemukan di Pricelist.")


def find_header_row_by_candidates(
    ws: Worksheet,
    required_candidates: Dict[str, List[str]],
    scan_rows: int = 10
) -> Tuple[int, Dict[str, int]]:
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        found: Dict[str, int] = {}
        ok = True
        for key, candidates in required_candidates.items():
            col = get_header_col_fuzzy(ws, r, candidates)
            if col is None:
                ok = False
                break
            found[key] = col
        if ok:
            return r, found
    raise ValueError("Header tidak ditemukan. Pastikan file memiliki kolom yang dibutuhkan.")


# ============================================================
# STOCK PRICELIST HELPERS
# ============================================================
def sheet_range_between(sheetnames: List[str], start: str, end: str) -> List[str]:
    up = [su(x) for x in sheetnames]
    if start.upper() not in up or end.upper() not in up:
        raise ValueError(f"Sheet range tidak valid. Pastikan ada '{start}' dan '{end}'.")
    i0 = up.index(start.upper())
    i1 = up.index(end.upper())
    if i0 > i1:
        i0, i1 = i1, i0
    return sheetnames[i0:i1 + 1]


def delete_coming_block_in_laptop(ws: Worksheet):
    r_start = find_row_contains(ws, "COMING", scan_rows=600)
    r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
    if r_start and r_end and r_end >= r_start:
        ws.delete_rows(r_start, r_end - r_start + 1)


def find_tot_col(ws: Worksheet, header_row_hint: int) -> Tuple[int, int]:
    for c in range(1, ws.max_column + 1):
        if su(ws.cell(header_row_hint, c).value) == "TOT":
            return header_row_hint, c
    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == "TOT":
                return r, c
    raise ValueError("Kolom 'TOT' tidak ketemu.")


def build_area_warehouse_meta(
    ws: Worksheet,
    merged_map: Dict[Tuple[int, int], object],
    area_row: int,
    warehouse_row: int,
    start_col: int,
) -> Dict[int, Dict[str, str]]:
    col_area_wh: Dict[int, Dict[str, str]] = {}
    for c in range(start_col, ws.max_column + 1):
        area_raw = get_cell_or_merged_value(ws, merged_map, area_row, c)
        warehouse_raw = get_cell_or_merged_value(ws, merged_map, warehouse_row, c)
        area_name = su(area_raw)
        warehouse_name = su(warehouse_raw)
        if not area_name or not warehouse_name:
            continue
        col_area_wh[c] = {
            "area": area_name,
            "warehouse": warehouse_name,
            "area_wh": f"{area_name}-{warehouse_name}",
        }
    return col_area_wh


def build_stock_lookup_from_sheet_fast(ws: Worksheet, sheet_name: str):
    header_row = find_header_row_by_exact(ws, "KODEBARANG", scan_rows=200)
    if header_row is None:
        header_row = find_header_row_by_exact(ws, "KODE BARANG", scan_rows=200)
    if header_row is None:
        raise ValueError(f"[{sheet_name}] Header 'KODEBARANG' tidak ketemu.")

    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v in ("KODEBARANG", "KODE BARANG"):
            sku_col = c
            break
    if sku_col is None:
        raise ValueError(f"[{sheet_name}] Kolom 'KODEBARANG' / 'KODE BARANG' tidak ditemukan.")

    header_row_used, tot_col = find_tot_col(ws, header_row)
    merged_map = build_merged_lookup_map(ws)
    area_row = header_row_used + 1
    warehouse_row = header_row_used + 2
    col_area_wh = build_area_warehouse_meta(
        ws,
        merged_map,
        area_row=area_row,
        warehouse_row=warehouse_row,
        start_col=tot_col + 1,
    )

    sku_map: Dict[str, Dict[str, Any]] = {}
    area_names: Set[str] = set()
    area_warehouses: Set[str] = set()
    for meta in col_area_wh.values():
        area_names.add(meta["area"])
        area_warehouses.add(meta["area_wh"])

    for r in range(max(header_row, warehouse_row) + 1, ws.max_row + 1):
        sku = s_clean(ws.cell(r, sku_col).value)
        if not sku:
            continue
        sku_key = norm_sku(sku)
        if sku_key in ("TOTAL", "KODEBARANG", "KODE BARANG", "KODEBARANG."):
            continue

        tot_val = to_int_or_none(ws.cell(r, tot_col).value)
        by_area_wh: Dict[str, int] = {}
        by_area: Dict[str, int] = {}
        for c, meta in col_area_wh.items():
            v = to_int_or_none(ws.cell(r, c).value)
            if v is None:
                continue
            area_wh_name = meta["area_wh"]
            area_name = meta["area"]
            by_area_wh[area_wh_name] = by_area_wh.get(area_wh_name, 0) + int(v)
            by_area[area_name] = by_area.get(area_name, 0) + int(v)
        sku_map[sku_key] = {"TOT": tot_val, "by_area_wh": by_area_wh, "by_area": by_area}
    return sku_map, {"area_options": sorted(area_names), "gudang_options": sorted(area_warehouses), "default_gudang_options": sorted(area_warehouses)}


def build_stock_lookup_from_pricelist_bytes(pl_bytes: bytes):
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True, read_only=False)
    for sname in wb.sheetnames:
        if su(sname) == "LAPTOP":
            delete_coming_block_in_laptop(wb[sname])
            break

    requested = {su(name): name for name in STOCK_PRICELIST_SHEETS}
    target_sheets = [sname for sname in wb.sheetnames if su(sname) in requested]
    if not target_sheets:
        raise ValueError(
            f"Sheet stok tidak ditemukan. Pricelist harus punya minimal salah satu sheet: {', '.join(STOCK_PRICELIST_SHEETS)}"
        )

    merged_lookup: Dict[str, Dict[str, Any]] = {}
    area_options_all: Set[str] = set()
    gudang_options_all: Set[str] = set()
    default_gudang_options_all: Set[str] = set()
    for sname in target_sheets:
        sku_map, meta = build_stock_lookup_from_sheet_fast(wb[sname], sname)
        merged_lookup.update(sku_map)
        area_options_all |= set(meta.get("area_options", []))
        gudang_options_all |= set(meta.get("gudang_options", []))
        default_gudang_options_all |= set(meta.get("default_gudang_options", []))
    if not merged_lookup:
        raise ValueError("Pricelist terbaca, tapi lookup stok kosong.")
    return merged_lookup, {
        "area_options": sorted(area_options_all),
        "gudang_options": sorted(gudang_options_all),
        "default_gudang_options": sorted(default_gudang_options_all),
    }


def apply_stock_floor_rule(qty: Optional[int], zero_below: int = 0) -> Optional[int]:
    if qty is None:
        return None
    qty = int(qty)
    if zero_below > 0 and qty < int(zero_below):
        return 0
    return qty


def get_default_tongle_gudangs(area_warehouses: List[str]) -> List[str]:
    area_wh_set = {su(a) for a in area_warehouses}
    return [area for area in DEFAULT_TONGLE_GUDANGS if su(area) in area_wh_set]


def pick_stock_value(
    sku_full: str,
    stock_lookup: Dict[str, Dict],
    selected_modes: Set[str],
    chosen_areas: Set[str],
    chosen_gudangs: Set[str],
    zero_below: int = 0,
) -> Optional[int]:
    base, _ = split_sku_addons(sku_full)
    base_key = norm_sku(base)
    if not base_key or base_key not in stock_lookup:
        return None

    rec = stock_lookup[base_key]
    tot = rec.get("TOT")
    by_area = rec.get("by_area", {}) or {}
    by_area_wh = rec.get("by_area_wh", {}) or {}

    if "Stok Nasional (TOT)" in selected_modes:
        return apply_stock_floor_rule(tot if tot is not None else None, zero_below)

    picked_area_whs: Set[str] = set()
    if "Default" in selected_modes:
        picked_area_whs |= {a for a in DEFAULT_TONGLE_GUDANGS if a in by_area_wh}

    total = 0
    counted_area_whs: Set[str] = set()

    if "Area" in selected_modes:
        for area_name in chosen_areas:
            total += int(by_area.get(area_name, 0) or 0)
            counted_area_whs |= {k for k in by_area_wh.keys() if k.startswith(f"{area_name}-")}

    if "Gudang" in selected_modes:
        picked_area_whs |= {a for a in chosen_gudangs if a in by_area_wh}

    for area_wh in picked_area_whs:
        if area_wh in counted_area_whs:
            continue
        total += int(by_area_wh.get(area_wh, 0) or 0)

    if not selected_modes or total == 0 and not counted_area_whs and not picked_area_whs:
        return None

    return apply_stock_floor_rule(total, zero_below)


# ============================================================
# STOCK PROCESSORS
# ============================================================
def find_shopee_columns_readonly(ws) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    sku_col = None
    qty_col = None
    for idx, val in enumerate(row_vals, start=1):
        v = su(val)
        if v == "SKU":
            sku_col = idx
        if v == "STOK":
            qty_col = idx
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def find_shopee_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    sku_col = None
    qty_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v == "SKU":
            sku_col = c
        if v == "STOK":
            qty_col = c
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_shopee(file_bytes: bytes, stock_lookup: Dict[str, Dict], selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_shopee_columns_readonly(ws)

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_shopee_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_shopee_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_shopee_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        try:
            rows, stats = collect_changed_rows_stock_shopee(mf.getvalue(), stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
            changed_rows_all.extend(rows)
            for k in ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"):
                summary[k] += stats[k]
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    result_bytes = write_stock_shopee_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, make_issues_workbook(issues) if issues else None, summary


def find_tiktokshop_columns_readonly(ws) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 6
    qty_headers = {"KUANTITAS", "JUMLAH DI SHOP LOCATION", "QUANTITY"}
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    sku_col = None
    qty_col = None
    for idx, val in enumerate(row_vals, start=1):
        v = su(val)
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = idx
        if v in qty_headers:
            qty_col = idx
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def find_tiktokshop_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 6
    qty_headers = {"KUANTITAS", "JUMLAH DI SHOP LOCATION", "QUANTITY"}
    sku_col = None
    qty_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v in ("SKU PENJUAL", "SELLER SKU"):
            sku_col = c
        if v in qty_headers:
            qty_col = c
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_tiktokshop(file_bytes: bytes, stock_lookup: Dict[str, Dict], selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_tiktokshop_columns_readonly(ws)

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_tiktokshop_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_tiktokshop_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_tiktokshop_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        try:
            rows, stats = collect_changed_rows_stock_tiktokshop(mf.getvalue(), stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
            changed_rows_all.extend(rows)
            for k in ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"):
                summary[k] += stats[k]
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    result_bytes = write_stock_tiktokshop_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, make_issues_workbook(issues) if issues else None, summary


def find_bigseller_stock_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row, found_cols = find_header_row_by_candidates(
        ws,
        {
            "sku": ["SKU", "Seller SKU", "SKU Penjual"],
            "qty": ["Stock", "Stok", "Quantity", "Qty"],
        },
        scan_rows=10,
    )
    return header_row + 1, found_cols["sku"], found_cols["qty"]


def process_bigseller_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    output_parts: List[Tuple[str, bytes]] = []
    current_rows: List[List[Any]] = []
    current_part = 1
    output_header: List[Any] = []
    header_len = 0

    def flush_part():
        nonlocal current_rows, current_part, output_parts, output_header, header_len
        if not current_rows:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, val in enumerate(output_header, start=1):
            ws.cell(row=1, column=c).value = val
        for r_idx, row_vals in enumerate(current_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c).value = val
        output_parts.append((f"hasil_update_stok_bigseller_part_{current_part}.xlsx", workbook_to_bytes(wb)))
        current_rows = []
        current_part += 1

    for mf in mass_files:
        wb = None
        try:
            wb = load_workbook(io.BytesIO(mf.getvalue()), read_only=False, data_only=False)
            ws = wb.worksheets[0]
            data_start, sku_col, qty_col = find_bigseller_stock_columns(ws)

            if not output_header:
                header_row = data_start - 1
                output_header = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
                header_len = ws.max_column

            for r in range(data_start, ws.max_row + 1):
                sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue

                summary["rows_scanned"] += 1
                old_qty = to_int_or_none(ws.cell(row=r, column=qty_col).value)
                new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)

                if new_qty is None:
                    summary["rows_unmatched"] += 1
                    issues.append({
                        "file": mf.name,
                        "row": r,
                        "sku_full": sku_full,
                        "old_value": old_qty,
                        "new_value": "",
                        "reason": "SKU tidak ditemukan di Pricelist stok",
                    })
                    continue

                if old_qty is not None and int(old_qty) == int(new_qty):
                    continue

                row_vals = [ws.cell(row=r, column=c).value for c in range(1, header_len + 1)]
                row_vals[qty_col - 1] = int(new_qty)
                current_rows.append(row_vals)
                summary["rows_written"] += 1

                if len(current_rows) >= BIGSELLER_MAX_ROWS_PER_FILE:
                    flush_part()

        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    flush_part()
    summary["issues_count"] = len(issues)
    if not output_parts:
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        if output_header:
            for c, val in enumerate(output_header, start=1):
                empty_ws.cell(row=1, column=c).value = val
        output_parts.append(("hasil_update_stok_bigseller_part_1.xlsx", workbook_to_bytes(empty_wb)))
    if len(output_parts) == 1:
        return output_parts[0][1], output_parts[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_parts), "hasil_update_stok_bigseller.zip", make_issues_workbook(issues) if issues else None, summary




def find_blibli_stock_columns(ws: Worksheet) -> Tuple[int, int, int, int]:
    header_row = 1
    data_start = 5
    sku_col = get_header_col_fuzzy(ws, header_row, ["Seller SKU"])
    qty_col = get_header_col_fuzzy(ws, header_row, ["Stok", "Stock"])
    sheet_col = None
    for idx, sname in enumerate(ws.parent.sheetnames):
        if ws.parent[sname] is ws:
            sheet_col = idx
            break
    if sku_col is None or qty_col is None:
        raise ValueError("Kolom Seller SKU/Stok tidak ketemu pada template Blibli.")
    return data_start, sku_col, qty_col, 0


def process_blibli_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        data_start, sku_col, qty_col, _ = find_blibli_stock_columns(ws)

        changed_rows: List[int] = []
        for r in range(data_start, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_qty = to_int_or_none(ws.cell(row=r, column=qty_col).value)
            new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
            if new_qty is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_qty, "new_value": "", "reason": "SKU tidak ditemukan di Pricelist stok"})
                continue
            if old_qty is not None and int(old_qty) == int(new_qty):
                continue
            safe_set_cell_value(ws, r, qty_col, int(new_qty))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_update_stok_blibli_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_update_stok_blibli.zip", make_issues_workbook(issues) if issues else None, summary


def find_akulaku_stock_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 1
    data_start = 2
    sku_col = get_header_col_fuzzy(ws, header_row, ["SKU Produk"])
    qty_col = get_header_col_fuzzy(ws, header_row, ["Stok", "Stock"])
    if sku_col is None or qty_col is None:
        raise ValueError("Kolom SKU Produk/Stok tidak ketemu pada template Akulaku.")
    return data_start, sku_col, qty_col


def process_akulaku_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0):
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        data_start, sku_col, qty_col = find_akulaku_stock_columns(ws)

        changed_rows: List[int] = []
        for r in range(data_start, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_qty = to_int_or_none(ws.cell(row=r, column=qty_col).value)
            new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below)
            if new_qty is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_qty, "new_value": "", "reason": "SKU tidak ditemukan di Pricelist stok"})
                continue
            if old_qty is not None and int(old_qty) == int(new_qty):
                continue
            safe_set_cell_value(ws, r, qty_col, int(new_qty))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_update_stok_akulaku_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_update_stok_akulaku.zip", make_issues_workbook(issues) if issues else None, summary

# ============================================================
# PRICE LOADERS
# ============================================================
def load_addon_map_generic(addon_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active
    code_candidates = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON", "Standarisasi Kode SKU di Varian"]
    price_candidates = ["harga", "HARGA", "Price", "PRICE", "Harga"]

    header_row = None
    code_col = None
    price_col = None
    for r in range(1, 30):
        code_col = get_header_col_fuzzy(ws, r, code_candidates)
        price_col = get_header_col_fuzzy(ws, r, price_candidates)
        if code_col and price_col:
            header_row = r
            break
    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    addon_map: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue
        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue
        addon_map[code] = int(apply_multiplier_if_needed(price_raw))
    return addon_map


def find_header_row_and_cols_pricelist_fixed(ws: Worksheet, required_price_cols: List[str]) -> Tuple[int, int, Dict[str, int]]:
    sku_candidates = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]

    for header_row in range(1, min(25, ws.max_row) + 1):
        sku_col = get_header_col_fuzzy(ws, header_row, sku_candidates)
        if sku_col is None:
            continue

        price_cols: Dict[str, int] = {}
        all_found = True
        for p in required_price_cols:
            col = get_header_col_fuzzy(ws, header_row, [p])
            if col is None:
                all_found = False
                break
            price_cols[p] = col

        if all_found:
            return header_row, sku_col, price_cols

    raise ValueError(
        f"Header Pricelist tidak ketemu. Pastikan ada kolom SKU/KODEBARANG dan kolom harga {required_price_cols}."
    )


def load_pricelist_price_map(pl_bytes: bytes, needed_cols: List[str]) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = get_change_sheet(wb)
    header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, needed_cols)
    result: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku = norm_sku(ws.cell(row=r, column=sku_col).value)
        if not sku:
            continue
        result[sku] = {}
        for label, col in price_cols.items():
            raw = parse_price_cell(ws.cell(row=r, column=col).value)
            if raw is not None:
                result[sku][label] = int(apply_multiplier_if_needed(raw))
    return result


def load_pricelist_price_map_multisheet(
    pl_bytes: bytes,
    needed_cols: List[str],
    start_sheet: str = "LAPTOP",
    end_sheet: str = "ACC",
) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)

    for sname in wb.sheetnames:
        if su(sname) == "LAPTOP":
            delete_coming_block_in_laptop(wb[sname])
            break

    target_sheets = sheet_range_between(wb.sheetnames, start_sheet, end_sheet)
    result: Dict[str, Dict[str, int]] = {}
    parsed_sheets = 0

    for sname in target_sheets:
        ws = wb[sname]
        try:
            header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, needed_cols)
        except Exception:
            continue

        parsed_sheets += 1
        for r in range(header_row + 1, ws.max_row + 1):
            sku = norm_sku(ws.cell(row=r, column=sku_col).value)
            if not sku:
                continue
            if sku not in result:
                result[sku] = {}
            for label, col in price_cols.items():
                raw = parse_price_cell(ws.cell(row=r, column=col).value)
                if raw is not None:
                    result[sku][label] = int(apply_multiplier_if_needed(raw))

    if parsed_sheets == 0:
        raise ValueError(
            f"Tidak ada sheet pricelist yang valid pada range '{start_sheet}' s/d '{end_sheet}' untuk kolom {needed_cols}."
        )
    if not result:
        raise ValueError(
            f"Pricelist multi-sheet '{start_sheet}' s/d '{end_sheet}' terbaca, tapi data harga kosong."
        )
    return result


def compute_price_from_maps(sku_full: str, price_map: Dict[str, Dict[str, int]], addon_map: Dict[str, int], price_key: str, discount_rp: int) -> Tuple[Optional[int], str]:
    base_sku, addons = split_sku_addons(sku_full)
    base_sku = norm_sku(base_sku)
    if not base_sku:
        return None, "SKU kosong"
    pl = price_map.get(base_sku)
    if not pl:
        return None, f"Base SKU '{base_sku}' tidak ada di Pricelist"
    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist untuk SKU '{base_sku}'"
    addon_total = 0
    for addon in addons:
        code = normalize_addon_code(addon)
        if code and code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map.get(code, 0))
    final_price = int(base_price) + addon_total - int(discount_rp)
    if final_price <= 0:
        return None, f"Harga hasil {price_key} <= 0 untuk SKU '{base_sku}'"
    return final_price, f"{price_key} + addon - diskon"


# ============================================================
# PRICE PROCESSORS
# ============================================================
def _process_shopee_price_common(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str, page_title: str, mode: str):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        if mode == "normal":
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga", "Harga Diskon", "Harga diskon"],
                },
                scan_rows=10,
            )
        else:
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga Diskon", "Harga diskon", "Harga"],
                },
                scan_rows=10,
            )

        sku_col = found_cols["sku"]
        price_col = found_cols["price"]
        data_start_fixed = header_row + 1

        changed_rows: List[int] = []
        for r in range(data_start_fixed, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)

            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({
                    "file": mf.name,
                    "row": r,
                    "sku_full": sku_full,
                    "old_value": old_price,
                    "new_value": "",
                    "reason": reason,
                })
                continue

            if old_price is not None and int(old_price) == int(new_price):
                continue

            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start_fixed - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), f"hasil_{page_title.lower().replace(' ', '_')}.zip", make_issues_workbook(issues) if issues else None, summary




# PRICE PROCESSORS - GROUPED BY TYPE
# ============================================================
# Normal Price
def process_shopee_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    return _process_shopee_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "M4",
        "Harga Normal Shopee",
        "normal",
    )


def process_tiktokshop_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        sku_col = get_header_col_fuzzy(ws, 3, ["SKU Penjual", "Seller SKU"])
        price_col = get_header_col_fuzzy(ws, 3, ["Harga Ritel (Mata Uang Lokal)", "Harga", "Price"])
        if sku_col is None or price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update TikTokShop tidak sesuai."})
            continue

        changed_rows: List[int] = []
        for r in range(6, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M3", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, 5, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_harga_normal_tiktokshop_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_harga_normal_tiktokshop.zip", make_issues_workbook(issues) if issues else None, summary


def _process_powemerchant_price_common(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, page_title: str):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        sku_col = get_header_col_fuzzy(ws, 3, ["SKU Penjual", "Seller SKU"])
        price_col = get_header_col_fuzzy(ws, 3, ["Harga Ritel (Mata Uang Lokal)", "Harga", "Price"])
        if sku_col is None or price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update PowerMerchant tidak sesuai."})
            continue

        changed_rows: List[int] = []
        for r in range(6, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M4", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, 5, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), f"hasil_{page_title.lower().replace(' ', '_')}.zip", make_issues_workbook(issues) if issues else None, summary


def process_powemerchant_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    return _process_powemerchant_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "Harga Normal PowerMerchant",
    )


def process_bigseller_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    output_parts: List[Tuple[str, bytes]] = []
    current_rows: List[List[Any]] = []
    current_part = 1
    output_header: List[Any] = []
    header_len = 0

    def flush_part():
        nonlocal current_rows, current_part, output_parts, output_header, header_len
        if not current_rows:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, val in enumerate(output_header, start=1):
            ws.cell(row=1, column=c).value = val
        for r_idx, row_vals in enumerate(current_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c).value = val
        output_parts.append((f"hasil_harga_normal_bigseller_part_{current_part}.xlsx", workbook_to_bytes(wb)))
        current_rows = []
        current_part += 1

    for mf in mass_files:
        wb = None
        try:
            wb = load_workbook(io.BytesIO(mf.getvalue()), read_only=False, data_only=False)
            ws = wb.worksheets[0]

            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "Seller SKU", "SKU Penjual"],
                    "price": ["Price", "Harga", "Harga Jual"],
                },
                scan_rows=10,
            )

            sku_col = found_cols["sku"]
            harga_col = found_cols["price"]

            if not output_header:
                output_header = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
                header_len = ws.max_column

            for r in range(header_row + 1, ws.max_row + 1):
                sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue

                summary["rows_scanned"] += 1
                old_price = parse_price_cell(ws.cell(row=r, column=harga_col).value)
                new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)

                if new_price is None:
                    summary["rows_unmatched"] += 1
                    issues.append({
                        "file": mf.name,
                        "row": r,
                        "sku_full": sku_full,
                        "old_value": old_price,
                        "new_value": "",
                        "reason": reason,
                    })
                    continue

                if old_price is not None and int(old_price) == int(new_price):
                    continue

                row_vals = [ws.cell(row=r, column=c).value for c in range(1, header_len + 1)]
                row_vals[harga_col - 1] = new_price
                current_rows.append(row_vals)
                summary["rows_written"] += 1

                if len(current_rows) >= BIGSELLER_MAX_ROWS_PER_FILE:
                    flush_part()

        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    flush_part()
    summary["issues_count"] = len(issues)
    if not output_parts:
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        if output_header:
            for c, val in enumerate(output_header, start=1):
                empty_ws.cell(row=1, column=c).value = val
        output_parts.append(("hasil_harga_normal_bigseller_part_1.xlsx", workbook_to_bytes(empty_wb)))
    if len(output_parts) == 1:
        return output_parts[0][1], output_parts[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_parts), "hasil_harga_normal_bigseller.zip", make_issues_workbook(issues) if issues else None, summary




def find_blibli_price_columns(ws: Worksheet) -> Tuple[int, int, int, int]:
    header_row = 1
    data_start = 5
    sku_col = get_header_col_fuzzy(ws, header_row, ["Seller SKU"])
    price_col = get_header_col_fuzzy(ws, header_row, ["Harga (Rp)", "Harga"])
    sale_price_col = get_header_col_fuzzy(ws, header_row, ["Harga Penjualan (Rp)", "Harga Penjualan"])
    if sku_col is None or price_col is None or sale_price_col is None:
        raise ValueError("Kolom Seller SKU/Harga/Harga Penjualan tidak ketemu pada template Blibli.")
    return data_start, sku_col, price_col, sale_price_col


def process_blibli_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        data_start, sku_col, price_col, sale_price_col = find_blibli_price_columns(ws)

        changed_rows: List[int] = []
        for r in range(data_start, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            old_sale_price = parse_price_cell(ws.cell(row=r, column=sale_price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M3", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and old_sale_price is not None and int(old_price) == int(new_price) and int(old_sale_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            safe_set_cell_value(ws, r, sale_price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_harga_normal_blibli_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_harga_normal_blibli.zip", make_issues_workbook(issues) if issues else None, summary


def find_akulaku_price_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 1
    data_start = 2
    sku_col = get_header_col_fuzzy(ws, header_row, ["SKU Produk"])
    price_col = get_header_col_fuzzy(ws, header_row, ["Harga", "Price"])
    if sku_col is None or price_col is None:
        raise ValueError("Kolom SKU Produk/Harga tidak ketemu pada template Akulaku.")
    return data_start, sku_col, price_col


def process_akulaku_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {"files_total": len(mass_files), "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}

    for mf in mass_files:
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        data_start, sku_col, price_col = find_akulaku_price_columns(ws)

        changed_rows: List[int] = []
        for r in range(data_start, ws.max_row + 1):
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M3", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            keep = set(changed_rows)
            for r in range(ws.max_row, data_start - 1, -1):
                if r not in keep:
                    ws.delete_rows(r, 1)
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        output_files.append((f"hasil_harga_normal_akulaku_{mf.name}", workbook_to_bytes(wb)))

    summary["issues_count"] = len(issues)
    if len(output_files) == 1:
        return output_files[0][1], output_files[0][0], make_issues_workbook(issues) if issues else None, summary
    return zip_named_files(output_files), "hasil_harga_normal_akulaku.zip", make_issues_workbook(issues) if issues else None, summary

# ============================================================
# SUBMIT CAMPAIGN PROCESSORS
# ============================================================

# Discount Price
def process_shopee_discount(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    return _process_shopee_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "M4",
        "Harga Coret Shopee",
        "coret",
    )


def process_tiktokshop_discount(input_file: Any, pricelist_file: Any, addon_file: Any, discount_rp: int, only_changed: bool = True):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    wb_in = load_workbook(io.BytesIO(input_file.getvalue()), data_only=True)
    ws_in = wb_in.active

    out_wb = Workbook()
    ws_out = out_wb.active
    ws_out.title = "Sheet1"
    headers = [
        "Product_id (wajib)",
        "SKU_id (wajib)",
        "Harga Penawaran (wajib)",
        "Total Stok Promosi (opsional)\n1. Total Stok Promosi≤ Stok \n2. Jika tidak diisi artinya tidak terbatas",
        "Batas Pembelian (opsional)\n1. 1 ≤ Batas pembelian≤ 99\n2. Jika tidak diisi artinya tidak terbatas",
    ]
    for i, h in enumerate(headers, start=1):
        ws_out.cell(row=1, column=i).value = h

    issues: List[Dict[str, Any]] = []
    summary = {"files_total": 1, "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    row_out = 2

    for r in range(6, ws_in.max_row + 1):
        product_id = parse_number_like_id(ws_in.cell(row=r, column=1).value)
        sku_id = parse_number_like_id(ws_in.cell(row=r, column=4).value)
        old_price = parse_price_cell(ws_in.cell(row=r, column=6).value)
        stock = to_int_or_none(ws_in.cell(row=r, column=7).value)
        seller_sku = s_clean(ws_in.cell(row=r, column=8).value or ws_in.cell(row=r, column=5).value)
        if not seller_sku:
            continue
        summary["rows_scanned"] += 1
        new_price, reason = compute_price_from_maps(seller_sku, price_map, addon_map, "M3", discount_rp)
        if new_price is None:
            summary["rows_unmatched"] += 1
            issues.append({"file": input_file.name, "row": r, "sku_full": seller_sku, "old_value": old_price, "new_value": "", "reason": reason})
            continue
        if only_changed and old_price is not None and int(old_price) == int(new_price):
            continue
        ws_out.cell(row=row_out, column=1).value = product_id
        ws_out.cell(row=row_out, column=2).value = sku_id
        ws_out.cell(row=row_out, column=3).value = new_price
        ws_out.cell(row=row_out, column=4).value = stock if stock is not None else ""
        row_out += 1
        summary["rows_written"] += 1

    summary["issues_count"] = len(issues)
    return workbook_to_bytes(out_wb), "hasil_harga_coret_tiktokshop.xlsx", make_issues_workbook(issues) if issues else None, summary


def process_powemerchant_discount(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int):
    return _process_powemerchant_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "Harga Coret PowerMerchant",
    )




def process_tiktokshop_campaign(mass_files: List[Any]):
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {
        "files_total": len(mass_files),
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unmatched": 0,
        "issues_count": 0,
    }

    for mf in mass_files:
        src_wb = load_workbook(io.BytesIO(mf.getvalue()))
        src_ws = src_wb.active

        header_row = 2
        data_start = 3

        sku_col = get_header_col_fuzzy(src_ws, header_row, [
            "SKU Name",
            "Nama SKU",
            "Seller SKU",
            "SKU Penjual",
        ])

        if sku_col is None:
            issues.append({
                "file": mf.name,
                "reason": "Header Submit Campaign tidak sesuai. Pastikan row 2 berisi header SKU Name.",
            })
            continue

        filtered_rows_data: List[List[Any]] = []

        for r in range(data_start, src_ws.max_row + 1):
            sku_full = s_clean(src_ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            if "ND-ALL-CAMPAIGN" in su(sku_full):
                row_vals = [src_ws.cell(row=r, column=c).value for c in range(1, src_ws.max_column + 1)]
                filtered_rows_data.append(row_vals)
                summary["rows_written"] += 1
            else:
                summary["rows_unmatched"] += 1

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = src_ws.title

        for c in range(1, src_ws.max_column + 1):
            out_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)
            out_ws.cell(row=2, column=c, value=src_ws.cell(row=2, column=c).value)

        if filtered_rows_data:
            for out_r, row_vals in enumerate(filtered_rows_data, start=data_start):
                for c, val in enumerate(row_vals, start=1):
                    out_ws.cell(row=out_r, column=c, value=val)
        else:
            issues.append({
                "file": mf.name,
                "reason": "Tidak ada baris dengan SKU Name yang mengandung 'ND-ALL-CAMPAIGN'.",
            })

        output_files.append((f"hasil_submit_campaign_tiktokshop_{mf.name}", workbook_to_bytes(out_wb)))

    summary["issues_count"] = len(issues)

    if len(output_files) == 1:
        return (
            output_files[0][1],
            output_files[0][0],
            make_issues_workbook(issues) if issues else None,
            summary,
        )

    return (
        zip_named_files(output_files),
        "hasil_submit_campaign_tiktokshop.zip",
        make_issues_workbook(issues) if issues else None,
        summary,
    )


# ============================================================
# UI HELPERS
# ============================================================
def page_header(title: str, desc: str, requirements: List[str]):
    st.title(title)
    st.caption(desc)
    with st.expander("Kebutuhan File", expanded=True):
        for item in requirements:
            st.write(f"- {item}")


def render_stock_controls(area_key_prefix: str, pricelist_file: Any, mode_key: str, loaded_areas_key: str, load_button_key: str):
    selected_modes = set(st.multiselect(
        "Mode Stok",
        ["Stok Nasional (TOT)", "Default", "Area", "Gudang"],
        default=["Default"],
        key=mode_key,
        help="Selain Stok Nasional (TOT), mode stok bisa dipilih lebih dari 1. Jika pilih TOT bersamaan dengan mode lain, hasil stok akan pakai TOT.",
    ))
    zero_below = st.number_input("Stok < angka ini jadi 0", min_value=0, value=0, step=1, key=f"{area_key_prefix}_zero_below")

    needs_lookup_data = bool(selected_modes & {"Default", "Area", "Gudang"})

    if st.button("Load Data Area / Gudang", key=load_button_key):
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
        else:
            try:
                _, meta = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue())
                st.session_state[loaded_areas_key] = meta
                default_gudangs = get_default_tongle_gudangs(meta.get("gudang_options", []))
                if default_gudangs:
                    st.session_state[f"{area_key_prefix}_gudangs"] = default_gudangs
                st.success(
                    f"Data berhasil dimuat: {len(meta.get('area_options', []))} area, {len(meta.get('gudang_options', []))} kombinasi area-gudang"
                )
            except Exception as e:
                st.error(f"Gagal load data area / gudang: {e}")

    meta = st.session_state.get(loaded_areas_key, {}) or {}
    areas = meta.get("area_options", [])
    gudangs = meta.get("gudang_options", [])
    default_gudangs = st.session_state.get(f"{area_key_prefix}_gudangs", get_default_tongle_gudangs(gudangs))

    if "Default" in selected_modes:
        st.caption("Mode Default mengunci gudang: JKT-1A, JKT-3B, JKT-3C, JKT-4B")

    chosen_areas: Set[str] = set()
    chosen_gudangs: Set[str] = set()
    if "Area" in selected_modes:
        chosen_areas = set(st.multiselect(
            "Pilih Area",
            areas,
            default=st.session_state.get(f"{area_key_prefix}_areas", []),
            key=f"{area_key_prefix}_areas",
        ))

    if "Gudang" in selected_modes:
        chosen_gudangs = set(st.multiselect(
            "Pilih Gudang (format Area-Gudang)",
            gudangs,
            default=default_gudangs if "Default" in selected_modes else st.session_state.get(f"{area_key_prefix}_gudangs", []),
            key=f"{area_key_prefix}_gudangs",
        ))

    process_disabled = False
    if not selected_modes:
        process_disabled = True
    elif needs_lookup_data and not (areas or gudangs):
        process_disabled = True
    elif "Area" in selected_modes and not chosen_areas:
        process_disabled = True
    elif "Gudang" in selected_modes and not chosen_gudangs:
        process_disabled = True

    return selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled


def validate_mass_uploads(mass_files: List[Any]) -> Optional[str]:
    if not mass_files:
        return "Upload file mass update minimal 1 file."
    if len(mass_files) > MAX_MASS_FILES:
        return f"Maksimal {MAX_MASS_FILES} file per proses."
    if total_upload_size_mb(mass_files) > MAX_TOTAL_UPLOAD_MB:
        return f"Total upload melebihi {MAX_TOTAL_UPLOAD_MB} MB."
    return None


def run_with_loading(process_fn, loading_text: str = "Memproses..."):
    progress = st.progress(0, text=loading_text)
    try:
        progress.progress(20, text=loading_text)
        result = process_fn()
        progress.progress(100, text="Selesai")
        return result
    finally:
        progress.empty()


# ============================================================
# PAGES
# ============================================================
def render_dashboard():
    st.title(APP_TITLE)
    st.markdown("Aplikasi all-in-one untuk **Update Stok**, **Harga Normal**, **Harga Coret**, dan **Submit Campaign** marketplace.")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    with c1:
        st.subheader("Update Stok")
        st.write("- Shopee (Mall & Star)\n- TikTokShop\n- Bigseller")
    with c2:
        st.subheader("Update Harga Normal")
        st.write("- Shopee (Mall & Star)\n- TikTokShop\n- PowerMerchant\n- Bigseller")
    with c3:
        st.subheader("Update Harga Coret")
        st.write("- Shopee (Mall & Star)\n- TikTokShop\n- PowerMerchant")
    with c4:
        st.subheader("Submit Campaign")
        st.write("- TikTokShop\n- Shopee (Coming Soon)")
    with c5:
        st.subheader("Analisa Penjualan")
        st.write("- Redwood module")
    with c6:
        st.subheader("Analisa Produk & Stok")
        st.write("- Blackwood module")
    st.info("Gunakan menu di sidebar untuk memilih fitur.")


def render_update_stok_shopee():
    page_header(
        "Update Stok Shopee (Mall & Star)",
        "Memproses file mass update Shopee (Mall & Star) berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Shopee (.xlsx, Unprotect dulu)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Shopee (Mall & Star)", type=["xlsx"], accept_multiple_files=True, key="stock_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_shopee_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled = render_stock_controls(
        area_key_prefix="stock_shopee",
        pricelist_file=pricelist_file,
        mode_key="stock_shopee_mode",
        loaded_areas_key="stock_shopee_areas_loaded",
        load_button_key="load_area_shopee",
    )

    if st.button("Proses", key="btn_stock_shopee", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = run_with_loading(
                lambda: process_shopee_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below),
                "Memproses update stok Shopee...",
            )
            cache_downloads("stock_shopee", "hasil_update_stok_shopee.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_shopee")
    render_downloads("stock_shopee")


def render_update_stok_tiktokshop():
    page_header(
        "Update Stok TikTokShop",
        "Memproses file mass update TikTokShop berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update TikTokShop (.xlsx, Unprotect dulu)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update TikTokShop", type=["xlsx"], accept_multiple_files=True, key="stock_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_tiktokshop_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled = render_stock_controls(
        area_key_prefix="stock_tiktokshop",
        pricelist_file=pricelist_file,
        mode_key="stock_tiktokshop_mode",
        loaded_areas_key="stock_tiktokshop_areas_loaded",
        load_button_key="load_area_tiktokshop",
    )

    if st.button("Proses", key="btn_stock_tiktokshop", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = run_with_loading(
                lambda: process_tiktokshop_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below),
                "Memproses update stok TikTokShop...",
            )
            cache_downloads("stock_tiktokshop", "hasil_update_stok_tiktokshop.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_tiktokshop")
    render_downloads("stock_tiktokshop")


def render_update_stok_bigseller():
    page_header(
        "Update Stok Bigseller",
        "Mengubah stok Bigseller berdasarkan sheet pricelist LAPTOP, TELCO, dan PC HOM ELE. Output hanya baris yang berubah dan otomatis split 10.000 row per file.",
        [
            "Mass Update Bigseller (.xlsx, bisa banyak)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Bigseller", type=["xlsx"], accept_multiple_files=True, key="stock_bigseller_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_bigseller_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled = render_stock_controls(
        area_key_prefix="stock_bigseller",
        pricelist_file=pricelist_file,
        mode_key="stock_bigseller_mode",
        loaded_areas_key="stock_bigseller_areas_loaded",
        load_button_key="load_area_bigseller",
    )

    if st.button("Proses", key="btn_stock_bigseller", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_bigseller_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below),
                "Memproses update stok Bigseller...",
            )
            cache_downloads("stock_bigseller", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_bigseller")
    render_downloads("stock_bigseller")




def render_update_stok_blibli():
    page_header(
        "Update Stok Blibli",
        "Memproses file mass update Blibli berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Blibli (.xlsx)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Blibli", type=["xlsx"], accept_multiple_files=True, key="stock_blibli_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_blibli_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled = render_stock_controls(
        area_key_prefix="stock_blibli",
        pricelist_file=pricelist_file,
        mode_key="stock_blibli_mode",
        loaded_areas_key="stock_blibli_areas_loaded",
        load_button_key="load_area_blibli",
    )

    if st.button("Proses", key="btn_stock_blibli", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_blibli_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below),
                "Memproses update stok Blibli...",
            )
            cache_downloads("stock_blibli", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_blibli")
    render_downloads("stock_blibli")


def render_update_stok_akulaku():
    page_header(
        "Update Stok Akulaku",
        "Memproses file mass update Akulaku berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Akulaku (.xlsx)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Akulaku", type=["xlsx"], accept_multiple_files=True, key="stock_akulaku_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_akulaku_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, process_disabled = render_stock_controls(
        area_key_prefix="stock_akulaku",
        pricelist_file=pricelist_file,
        mode_key="stock_akulaku_mode",
        loaded_areas_key="stock_akulaku_areas_loaded",
        load_button_key="load_area_akulaku",
    )

    if st.button("Proses", key="btn_stock_akulaku", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_akulaku_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below),
                "Memproses update stok Akulaku...",
            )
            cache_downloads("stock_akulaku", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_akulaku")
    render_downloads("stock_akulaku")


def render_harga_normal_shopee():
    page_header(
        "Harga Normal Shopee (Mall & Star)",
        "Mengubah harga normal Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update Shopee (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_shopee_disc")

    if st.button("Proses", key="btn_normal_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_shopee_price(
                    mass_files, pricelist_file, addon_file, discount_rp
                ),
                "Memproses harga normal Shopee...",
            )
            cache_downloads("normal_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_shopee")
    render_downloads("normal_shopee")




def render_harga_normal_blibli():
    page_header(
        "Harga Normal Blibli",
        "Mengubah harga normal Blibli berdasarkan sheet CHANGE di pricelist dan addon mapping. Kolom Harga (Rp) dan Harga Penjualan (Rp) akan diisi harga M3.",
        ["Template Mass Update Blibli (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_blibli_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_blibli_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_blibli_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_blibli_disc")

    if st.button("Proses", key="btn_normal_blibli"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_blibli_price(mass_files, pricelist_file, addon_file, discount_rp),
                "Memproses harga normal Blibli...",
            )
            cache_downloads("normal_blibli", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_blibli")
    render_downloads("normal_blibli")


def render_harga_normal_akulaku():
    page_header(
        "Harga Normal Akulaku",
        "Mengubah harga normal Akulaku berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update Akulaku (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_akulaku_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_akulaku_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_akulaku_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_akulaku_disc")

    if st.button("Proses", key="btn_normal_akulaku"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_akulaku_price(mass_files, pricelist_file, addon_file, discount_rp),
                "Memproses harga normal Akulaku...",
            )
            cache_downloads("normal_akulaku", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_akulaku")
    render_downloads("normal_akulaku")


def render_harga_coret_shopee():
    page_header(
        "Harga Coret Shopee (Mall & Star)",
        "Mengubah harga coret Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Discount Nominate Shopee (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Template Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_shopee_disc")

    if st.button("Proses", key="btn_coret_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_shopee_discount(
                    mass_files, pricelist_file, addon_file, discount_rp
                ),
                "Memproses harga coret Shopee...",
            )
            cache_downloads("coret_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_shopee")
    render_downloads("coret_shopee")


def render_harga_normal_tiktokshop():
    page_header(
        "Harga Normal TikTokShop",
        "Mengubah harga normal TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update TikTokShop (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_tiktokshop_disc")

    if st.button("Proses", key="btn_normal_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_tiktokshop_price(
                    mass_files, pricelist_file, addon_file, discount_rp
                ),
                "Memproses harga normal TikTokShop...",
            )
            cache_downloads("normal_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_tiktokshop")
    render_downloads("normal_tiktokshop")


def render_harga_coret_tiktokshop():
    page_header(
        "Harga Coret TikTokShop",
        "Membuat template output promo TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Input File TikTokShop (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        input_file = st.file_uploader("Upload File TikTokShop", type=["xlsx"], key="coret_tiktokshop_input")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_tiktokshop_disc")

    if st.button("Proses", key="btn_coret_tiktokshop"):
        if not input_file or not pricelist_file or not addon_file:
            st.error("Upload semua file yang dibutuhkan dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_tiktokshop_discount(
                    input_file, pricelist_file, addon_file, discount_rp, True
                ),
                "Memproses harga coret TikTokShop...",
            )
            cache_downloads("coret_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_tiktokshop")
    render_downloads("coret_tiktokshop")


def render_harga_normal_powemerchant():
    page_header(
        "Harga Normal PowerMerchant",
        "Mengubah harga normal PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_pm_disc")

    if st.button("Proses", key="btn_normal_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_powemerchant_price(
                    mass_files, pricelist_file, addon_file, discount_rp
                ),
                "Memproses harga normal PowerMerchant...",
            )
            cache_downloads("normal_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_pm")
    render_downloads("normal_pm")


def render_harga_coret_powemerchant():
    page_header(
        "Harga Coret PowerMerchant",
        "Mengubah harga coret PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_pm_disc")

    if st.button("Proses", key="btn_coret_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_powemerchant_discount(
                    mass_files, pricelist_file, addon_file, discount_rp
                ),
                "Memproses harga coret PowerMerchant...",
            )
            cache_downloads("coret_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_pm")
    render_downloads("coret_pm")


def render_harga_normal_bigseller():
    page_header(
        "Harga Normal Bigseller",
        "Mengubah harga Bigseller, hanya output baris yang berubah, dan otomatis split 10.000 row per file.",
        ["Mass Update Bigseller (.xlsx, bisa banyak)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_bigseller_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_bigseller_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_bigseller_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_bigseller_disc")
    price_key = st.radio(
        "Ambil harga dari Pricelist",
        ["M3", "M4"],
        horizontal=True,
        key="normal_bigseller_price_key",
    )

    if st.button("Proses", key="btn_normal_bigseller"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_bigseller_price(
                    mass_files, pricelist_file, addon_file, discount_rp, price_key
                ),
                "Memproses harga normal Bigseller...",
            )
            cache_downloads("normal_bigseller", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_bigseller")
    render_downloads("normal_bigseller")


def render_submit_campaign_shopee():
    page_header(
        "Submit Campaign Shopee",
        "Fitur submit campaign untuk Shopee sedang disiapkan.",
        ["Coming Soon"],
    )
    st.info("Coming Soon")


def render_submit_campaign_tiktokshop():
    page_header(
        "Submit Campaign TikTokShop",
        "Cukup upload template campaign TikTokShop. Sistem hanya akan menyimpan baris yang memiliki karakter 'ND-ALL-CAMPAIGN' pada kolom SKU Name.",
        [
            "Template Campaign Tiktokshop (.xlsx)",
        ],
    )

    mass_files = st.file_uploader(
        "Upload Template Campaign Tiktokshop",
        type=["xlsx"],
        accept_multiple_files=True,
        key="submit_campaign_tiktokshop_mass",
    )

    if st.button("Proses", key="btn_submit_campaign_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return

        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading(
                lambda: process_tiktokshop_campaign(mass_files=mass_files),
                "Memfilter template campaign TikTokShop...",
            )
            cache_downloads(
                "submit_campaign_tiktokshop",
                result_name,
                result_bytes,
                issues_bytes,
                summary=summary,
            )
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("submit_campaign_tiktokshop")
    render_downloads("submit_campaign_tiktokshop")




# ============================================================
# EXTERNAL MODULE: render_analisa_penjualan
# ============================================================
def render_analisa_penjualan():
    # =========================
    # CSS (Light + Badge + Clip + SKU column widths + Smaller headers)
    # =========================
    st.markdown(
        """
    <style>
    .block-container { padding-top: 0.6rem; padding-bottom: 1.2rem; }
    section[data-testid="stSidebar"] .block-container { padding-top: 0.6rem; }

    /* Header (large -> small) */
    .header-wrap { display:flex; align-items:center; gap:12px; margin: 0.2rem 0 0.8rem 0; }
    .header-title { font-size: 22px; font-weight: 900; margin:0; line-height:1.1; color: #111827; }

    /* Make section headers smaller */
    h2 { font-size: 18px !important; }
    h3 { font-size: 16px !important; }

    /* Cards */
    .kpi-grid { display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; }

    .card {
      border: 1px solid rgba(17,24,39,0.08);
      background: #ffffff;
      border-radius: 14px;
      padding: 12px 12px 10px 12px;
      box-shadow: 0 10px 20px rgba(17,24,39,0.04);
    }
    .card-title { font-size: 12px; color: rgba(17,24,39,0.7); margin-bottom: 6px; }
    .card-value { font-size: 18px; font-weight: 900; line-height: 1.15; color: #111827; }
    .card-sub { font-size: 11px; margin-top: 6px; font-weight: 800; }

    .pos { color: #16a34a; }
    .neg { color: #dc2626; }
    .na  { color: #64748b; }

    hr { border: none; border-top: 1px solid rgba(17,24,39,0.10); margin: 14px 0; }

    /* Small headings (for long titles) */
    .small-h { font-size: 16px; font-weight: 900; margin: 0 0 6px 0; }
    .small-h .muted { color: rgba(17,24,39,0.6); font-weight: 800; }

    /* HTML table styling + CLIP */
    table {
      width: 100%;
      border-collapse: separate;
      border-spacing: 0;
      overflow: hidden;
      border-radius: 12px;
      border: 1px solid rgba(17,24,39,0.08);
      background: #fff;
      table-layout: fixed;
    }
    thead th {
      background: rgba(17,24,39,0.03);
      font-size: 12px;
      color: rgba(17,24,39,0.8);
      padding: 10px 10px;
      border-bottom: 1px solid rgba(17,24,39,0.08);
      text-align: left;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    tbody td {
      padding: 9px 10px;
      font-size: 12px;
      border-bottom: 1px solid rgba(17,24,39,0.06);
      color: #111827;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    tbody tr:hover td { background: rgba(225,29,46,0.04); }

    /* Growth badges: merah/hijau */
    .badge-pos, .badge-neg, .badge-na {
      display:inline-block;
      font-size: 11px;
      font-weight: 900;
      padding: 2px 8px;
      border-radius: 999px;
      white-space: nowrap;
    }
    .badge-pos { background: rgba(22,163,74,0.12); color:#16a34a; }
    .badge-neg { background: rgba(220,38,38,0.12); color:#dc2626; }
    .badge-na  { background: rgba(100,116,139,0.12); color:#64748b; }

    /* SKU table column widths (SPESIFIKASI wider, Growth wider so header & % visible) */
    table.sku-table th:nth-child(1), table.sku-table td:nth-child(1) { width: 50%; }
    table.sku-table th:nth-child(2), table.sku-table td:nth-child(2) { width: 14%; }
    table.sku-table th:nth-child(3), table.sku-table td:nth-child(3) { width: 14%; }
    table.sku-table th:nth-child(4), table.sku-table td:nth-child(4) { width: 10%; } /* Delta */
    table.sku-table th:nth-child(5), table.sku-table td:nth-child(5) { width: 12%; } /* Growth */
    </style>
    """,
        unsafe_allow_html=True,
    )

    DEFAULT_DATE_FORMAT_HINT = "Format TGL: dd-mm-yyyy / dd/mm/yyyy / yyyy-mm-dd"


    def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df


    def coerce_numeric_series(s: pd.Series) -> pd.Series:
        def to_num(x):
            if pd.isna(x):
                return np.nan
            if isinstance(x, (int, float, np.integer, np.floating)):
                return float(x)
            txt = str(x).strip()
            txt = re.sub(r"[^0-9\-\.,]", "", txt)
            if txt in ("", "-", ".", ","):
                return np.nan
            txt = txt.replace(" ", "").replace(".", "").replace(",", "")
            try:
                return float(txt)
            except Exception:
                return np.nan

        return s.map(to_num)


    def parse_tgl(df: pd.DataFrame, col: str = "TGL") -> pd.Series:
        s = df[col]
        if np.issubdtype(s.dtype, np.datetime64):
            return pd.to_datetime(s, errors="coerce").dt.date
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
        return parsed.dt.date


    def ensure_required_columns(df: pd.DataFrame) -> Tuple[bool, str]:
        required = [
            "STATUS", "TGL", "TRANSAKSI", "TEAM",
            "PRODUCT", "BRAND", "QTY", "JUMLAH",
            "SO NO",
            "COUNTRY",
            "SPESIFIKASI",
            "NAMA CUSTOMER",
            "OTO",
        ]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return False, f"Kolom wajib tidak ditemukan: {', '.join(missing)}"
        return True, ""


    def drop_total_rows(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        if "NO" in out.columns:
            out = out[~out["NO"].astype(str).str.strip().str.upper().eq("TOTAL")].copy()
        for c in ["STATUS", "TGL"]:
            if c in out.columns:
                out = out[~out[c].astype(str).str.strip().str.upper().eq("TOTAL")].copy()
        return out.dropna(how="all")


    def format_idr(x: float) -> str:
        if pd.isna(x):
            return "-"
        n = int(round(float(x)))
        s = f"{n:,}".replace(",", ".")
        return f"IDR {s}"


    def format_int_id(x: float) -> str:
        if pd.isna(x):
            return "-"
        return f"{int(round(float(x))):,}".replace(",", ".")


    def compact_number(x: float) -> str:
        if pd.isna(x):
            return ""
        x = float(x)
        ax = abs(x)
        if ax >= 1_000_000_000:
            return f"{x/1_000_000_000:.2f}B".replace(".", ",")
        if ax >= 1_000_000:
            return f"{x/1_000_000:.2f}M".replace(".", ",")
        if ax >= 1_000:
            return f"{x/1_000:.2f}K".replace(".", ",")
        return str(int(round(x)))


    def safe_growth_pct(this_val: float, last_val: float) -> Optional[float]:
        if last_val is None or pd.isna(last_val):
            return None
        last_val = float(last_val)
        if last_val == 0.0:
            return None
        return (float(this_val) - last_val) / last_val * 100.0


    def growth_label(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return "N/A"
        s = f"{g:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        return ("▲ " + s) if g >= 0 else ("▼ " + s)


    def growth_badge_html(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return '<span class="badge-na">N/A</span>'
        s = f"{g:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        if g >= 0:
            return f'<span class="badge-pos">▲ {s}</span>'
        return f'<span class="badge-neg">▼ {s}</span>'


    def kpi_delta_class(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return "na"
        return "pos" if g >= 0 else "neg"


    @dataclass
    class CleanData:
        df: pd.DataFrame
        date_min: date
        date_max: date


    @st.cache_data(show_spinner=False)
    def read_excel_cached(file_bytes: bytes, sheet_name: str, header_row_1based: int) -> pd.DataFrame:
        bio = io.BytesIO(file_bytes)
        header_idx = int(header_row_1based) - 1
        if sheet_name.strip() == "":
            return pd.read_excel(bio, header=header_idx)
        return pd.read_excel(bio, sheet_name=sheet_name.strip(), header=header_idx)


    @st.cache_data(show_spinner=False)
    def clean_sales_df_cached(df_raw: pd.DataFrame) -> CleanData:
        df = normalize_columns(df_raw)
        ok, msg = ensure_required_columns(df)
        if not ok:
            raise ValueError(msg)

        df = drop_total_rows(df)

        keep_cols = [
            "STATUS", "TGL", "TRANSAKSI", "TEAM",
            "PRODUCT", "BRAND", "QTY", "JUMLAH",
            "SO NO",
            "COUNTRY",
            "SPESIFIKASI",
            "NAMA CUSTOMER",
            "OTO",
            "AREA",
        ]
        df = df[keep_cols].copy()

        df["TGL"] = parse_tgl(df, "TGL")
        if df["TGL"].isna().any():
            bad = df[df["TGL"].isna()].head(8)
            raise ValueError(
                f"Ada TGL gagal diparse.\n\nContoh:\n{bad[['TGL']].to_string(index=False)}\n\n{DEFAULT_DATE_FORMAT_HINT}"
            )

        df["QTY"] = pd.to_numeric(coerce_numeric_series(df["QTY"]), errors="coerce").fillna(0.0)
        df["JUMLAH"] = pd.to_numeric(coerce_numeric_series(df["JUMLAH"]), errors="coerce").fillna(0.0)

        for c in [
            "STATUS", "TRANSAKSI", "TEAM", "PRODUCT", "BRAND",
            "SO NO", "COUNTRY", "SPESIFIKASI", "NAMA CUSTOMER", "OTO"
        ]:
            df[c] = df[c].astype(str).str.strip()

        df["ROW_TYPE"] = np.where(df["STATUS"].str.upper().str.contains("RETUR"), "RETUR", "SO_OUT")
        df["OTO_YES"] = df["OTO"].str.upper().eq("YES")
        df["PLATFORM"] = df["NAMA CUSTOMER"].astype(str).str.strip()

        df = df[df["STATUS"].str.strip().ne("")].copy()
        return CleanData(df=df, date_min=df["TGL"].min(), date_max=df["TGL"].max())


    @st.cache_data(show_spinner=False)
    def compute_kpis_cached(df: pd.DataFrame) -> Dict[str, float]:
        sales = float(df["JUMLAH"].sum())
        qty = float(df["QTY"].sum())

        so = df[df["ROW_TYPE"] == "SO_OUT"]
        orders = float(so["SO NO"].nunique())

        returns = float(len(df[df["ROW_TYPE"] == "RETUR"]))
        aov = sales / orders if orders else np.nan
        return {"sales": sales, "qty": qty, "orders": orders, "returns": returns, "aov": float(aov) if not pd.isna(aov) else np.nan}


    def get_week_start(d: date) -> date:
        return d - timedelta(days=d.weekday())


    def month_start(d: date) -> date:
        return date(d.year, d.month, 1)


    def prev_month_same_day(d: date) -> date:
        first = month_start(d)
        prev_last = first - timedelta(days=1)
        day = min(d.day, prev_last.day)
        return date(prev_last.year, prev_last.month, day)


    def slice_period(df: pd.DataFrame, start: date, end_inclusive: date) -> pd.DataFrame:
        return df[(df["TGL"] >= start) & (df["TGL"] <= end_inclusive)].copy()


    def build_period_frames(df_all: pd.DataFrame, mode: str, df_upload_a: pd.DataFrame, df_upload_b: pd.DataFrame):
        if mode == "UPLOAD":
            return df_upload_a, df_upload_b, "Periode A", "Periode B"

        anchor = df_all["TGL"].max()

        if mode == "WOW":
            this_start = get_week_start(anchor)
            this_end = anchor
            last_end = this_start - timedelta(days=1)
            last_start = get_week_start(last_end)
            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "Week Lalu",
                "Week Ini",
            )

        if mode == "MOM":
            this_start = month_start(anchor)
            this_end = anchor
            last_anchor = prev_month_same_day(anchor)
            last_start = month_start(last_anchor)
            last_month_mask = (df_all["TGL"] >= last_start) & (df_all["TGL"] < this_start)
            last_end = df_all.loc[last_month_mask, "TGL"].max() if last_month_mask.any() else last_start
            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "Bulan Lalu",
                "Bulan Ini",
            )

        if mode == "MTD":
            this_start = month_start(anchor)
            this_end = anchor
            last_anchor = prev_month_same_day(anchor)
            last_start = month_start(last_anchor)

            next_month_first = month_start(anchor)
            last_month_last_day = next_month_first - timedelta(days=1)
            last_end_candidate = date(last_month_last_day.year, last_month_last_day.month, min(anchor.day, last_month_last_day.day))

            last_month_mask = (df_all["TGL"] >= last_start) & (df_all["TGL"] < this_start)
            if last_month_mask.any():
                last_end_data = df_all.loc[last_month_mask, "TGL"].max()
                last_end = min(last_end_candidate, last_end_data)
            else:
                last_end = last_end_candidate

            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "MTD Bulan Lalu",
                "MTD Bulan Ini",
            )

        return df_upload_a, df_upload_b, "Periode Lalu", "Periode Ini"


    def options_for(df_all: pd.DataFrame, col: str) -> List[str]:
        return sorted([v for v in df_all[col].dropna().unique().tolist() if str(v).strip() != ""])


    def apply_multifilter(df: pd.DataFrame, col: str, selected: List[str]) -> pd.DataFrame:
        if not selected:
            return df
        return df[df[col].isin(selected)].copy()


    @st.cache_data(show_spinner=False)
    def top_table_cached(df_this: pd.DataFrame, df_last: pd.DataFrame, by_col: str, metric: str, top_n: int) -> pd.DataFrame:
        agg_this = df_this.groupby(by_col, as_index=False).agg(THIS=(metric, "sum"))
        agg_last = df_last.groupby(by_col, as_index=False).agg(LAST=(metric, "sum"))
        merged = agg_this.merge(agg_last, on=by_col, how="outer").fillna(0.0)
        merged["DELTA"] = merged["THIS"] - merged["LAST"]
        merged["GROWTH_NUM"] = merged.apply(lambda r: safe_growth_pct(r["THIS"], r["LAST"]), axis=1)
        merged = merged.sort_values("THIS", ascending=False).head(top_n)

        if metric == "JUMLAH":
            merged["Periode Ini"] = merged["THIS"].map(format_idr)
            merged["Periode Lalu"] = merged["LAST"].map(format_idr)
            merged["Delta"] = merged["DELTA"].map(format_idr)
        else:
            merged["Periode Ini"] = merged["THIS"].map(format_int_id)
            merged["Periode Lalu"] = merged["LAST"].map(format_int_id)
            merged["Delta"] = merged["DELTA"].map(format_int_id)

        merged["Growth"] = merged["GROWTH_NUM"].apply(growth_badge_html)
        merged = merged[[by_col, "Periode Ini", "Periode Lalu", "Delta", "Growth"]]
        return merged


    def render_html_table(df: pd.DataFrame, table_class: str = ""):
        html = df.to_html(escape=False, index=False)
        if table_class:
            html = html.replace("<table ", f'<table class="{table_class}" ', 1)
        st.markdown(html, unsafe_allow_html=True)


    def small_title(text: str, hint: str = ""):
        hint_html = f' <span class="muted">{hint}</span>' if hint else ""
        st.markdown(f'<div class="small-h">{text}{hint_html}</div>', unsafe_allow_html=True)


    def render_header():
        st.markdown(
            """
    <div class="header-wrap">
      <div>
        <div class="header-title">Analisa Penjualan</div>
      </div>
    </div>
    """,
            unsafe_allow_html=True,
        )


    def style_growth_pct_df(df_in: pd.DataFrame):
        df = df_in.copy()

        def color_growth(val):
            try:
                if pd.isna(val):
                    return "color: #64748b;"
                return "color: #dc2626; font-weight: 800;" if float(val) < 0 else "color: #16a34a; font-weight: 800;"
            except Exception:
                return "color: #64748b;"

        return (
            df.style
            .format({"Growth %": "{:.2f}%"})
            .applymap(color_growth, subset=["Growth %"])
        )


    # ===== NEW: Team Down % Table =====
    @st.cache_data(show_spinner=False)
    def team_down_ratio_table_cached(df_last: pd.DataFrame, df_this: pd.DataFrame) -> pd.DataFrame:
        last = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        this = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        t = last.merge(this, on="TEAM", how="outer").fillna(0.0)
        t["DELTA"] = t["QTY_INI"] - t["QTY_LALU"]

        # total team aktif: punya activity di salah satu periode
        active = t[(t["QTY_LALU"] > 0) | (t["QTY_INI"] > 0)].copy()
        total = int(len(active))
        turun = int((active["DELTA"] < 0).sum())
        naik = int((active["DELTA"] > 0).sum())
        flat = int((active["DELTA"] == 0).sum())
        pct_turun = (turun / total * 100.0) if total else 0.0

        out = pd.DataFrame(
            {
                "Total TEAM aktif": [total],
                "TEAM Turun": [turun],
                "TEAM Naik": [naik],
                "TEAM Tetap": [flat],
                "% TEAM Turun": [pct_turun],
            }
        )
        return out


    def _drivers_as_text(df_delta: pd.DataFrame, team_dir: str, top_k: int = 3) -> Dict[str, str]:
        """
        df_delta columns: TEAM, DIM, DELTA
        team_dir: mapping TEAM -> +1 (naik) or -1 (turun)
        """
        # Join direction for filtering
        d = df_delta.copy()
        d["DIR"] = d["TEAM"].map(team_dir).fillna(0).astype(int)

        # For naik: keep DELTA > 0, take top_k biggest
        # For turun: keep DELTA < 0, take top_k most negative
        naik_df = d[(d["DIR"] > 0) & (d["DELTA"] > 0)].copy()
        turun_df = d[(d["DIR"] < 0) & (d["DELTA"] < 0)].copy()

        naik_df = naik_df.sort_values(["TEAM", "DELTA"], ascending=[True, False]).groupby("TEAM").head(top_k)
        turun_df = turun_df.sort_values(["TEAM", "DELTA"], ascending=[True, True]).groupby("TEAM").head(top_k)

        # Build strings
        out: Dict[str, str] = {}

        def fmt_row(dim: str, delta: float) -> str:
            sign = "+" if delta > 0 else ""
            return f"{dim} ({sign}{int(delta):,})".replace(",", ".")

        for team, g in naik_df.groupby("TEAM"):
            out[team] = ", ".join([fmt_row(r["DIM"], r["DELTA"]) for _, r in g.iterrows()])

        for team, g in turun_df.groupby("TEAM"):
            out[team] = ", ".join([fmt_row(r["DIM"], r["DELTA"]) for _, r in g.iterrows()])

        return out


    @st.cache_data(show_spinner=False)
    def team_driver_analysis_table_cached(df_last: pd.DataFrame, df_this: pd.DataFrame, top_k: int = 3) -> pd.DataFrame:
        # team totals
        last_t = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        this_t = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        team = last_t.merge(this_t, on="TEAM", how="outer").fillna(0.0)
        team = team[(team["QTY_LALU"] > 0) | (team["QTY_INI"] > 0)].copy()
        team["DELTA_QTY"] = team["QTY_INI"] - team["QTY_LALU"]
        team["GROWTH_PCT"] = team.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        # direction mapping (+1 naik, -1 turun, 0 flat)
        team_dir = {r["TEAM"]: (1 if r["DELTA_QTY"] > 0 else (-1 if r["DELTA_QTY"] < 0 else 0)) for _, r in team.iterrows()}

        # build delta per TEAM x DIM for PRODUCT / BRAND / SKU(SPESIFIKASI)
        def build_delta(dim_col: str) -> pd.DataFrame:
            a = df_this.groupby(["TEAM", dim_col], as_index=False).agg(THIS=("QTY", "sum"))
            b = df_last.groupby(["TEAM", dim_col], as_index=False).agg(LAST=("QTY", "sum"))
            m = a.merge(b, on=["TEAM", dim_col], how="outer").fillna(0.0)
            m["DELTA"] = m["THIS"] - m["LAST"]
            m = m.rename(columns={dim_col: "DIM"})
            return m[["TEAM", "DIM", "DELTA"]]

        prod_delta = build_delta("PRODUCT")
        brand_delta = build_delta("BRAND")
        sku_delta = build_delta("SPESIFIKASI")

        prod_map = _drivers_as_text(prod_delta, team_dir, top_k=top_k)
        brand_map = _drivers_as_text(brand_delta, team_dir, top_k=top_k)
        sku_map = _drivers_as_text(sku_delta, team_dir, top_k=top_k)

        team["Arah"] = team["DELTA_QTY"].apply(lambda x: "NAIK" if x > 0 else ("TURUN" if x < 0 else "TETAP"))
        team["Produk (driver)"] = team["TEAM"].map(prod_map).fillna("-")
        team["Brand (driver)"] = team["TEAM"].map(brand_map).fillna("-")
        team["SKU/Spesifikasi (driver)"] = team["TEAM"].map(sku_map).fillna("-")

        # pretty
        out = team.copy()
        out["QTY Lalu"] = out["QTY_LALU"].map(format_int_id)
        out["QTY Ini"] = out["QTY_INI"].map(format_int_id)
        out["Delta"] = out["DELTA_QTY"].map(lambda x: f"{int(x):,}".replace(",", "."))
        out["Growth %"] = out["GROWTH_PCT"].apply(lambda x: float(x) if (x is not None and not pd.isna(x)) else np.nan)

        out = out.sort_values(["DELTA_QTY"], ascending=True)  # yang turun paling parah di atas (biar langsung keliatan)
        out = out[
            [
                "TEAM", "Arah", "QTY Lalu", "QTY Ini", "Delta", "Growth %",
                "Produk (driver)", "Brand (driver)", "SKU/Spesifikasi (driver)"
            ]
        ]
        return out


    # =========================
    # UI
    # =========================
    render_header()

    with st.sidebar:
        st.subheader("Upload Data")
        file_a = st.file_uploader("Excel A (.xlsx) — periode lama", type=["xlsx"], key="a")
        file_b = st.file_uploader("Excel B (.xlsx) — periode baru", type=["xlsx"], key="b")

        st.markdown("---")
        st.subheader("Header & Sheet")
        header_row_a = st.number_input("Header row Excel A (mulai dari 1)", 1, 30, 2, 1)
        header_row_b = st.number_input("Header row Excel B (mulai dari 1)", 1, 30, 2, 1)
        sheet_a = st.text_input("Nama sheet Excel A (kosongkan = sheet pertama)", "")
        sheet_b = st.text_input("Nama sheet Excel B (kosongkan = sheet pertama)", "")

    if not file_a or not file_b:
        st.info("Upload 2 file Excel dulu.")
        st.stop()

    with st.spinner("Membaca & membersihkan Excel (sekali di awal)..."):
        df_a_raw = read_excel_cached(file_a.getvalue(), sheet_a, header_row_a)
        df_b_raw = read_excel_cached(file_b.getvalue(), sheet_b, header_row_b)
        a = clean_sales_df_cached(df_a_raw)
        b = clean_sales_df_cached(df_b_raw)

    df_a = a.df
    df_b = b.df
    df_all = pd.concat([df_a, df_b], ignore_index=True)

    with st.sidebar:
        st.markdown("---")
        st.subheader("Mode Perbandingan")
        compare_mode = st.selectbox("Pilih periode", ["MOM", "WOW", "MTD", "UPLOAD"], 0)

        st.markdown("---")
        st.subheader("Opsi Tampilan")
        top_n = st.slider("Top N", 5, 30, 10, 1)
        metric_choice = st.selectbox("Metric", ["Qty (QTY)", "Sales (JUMLAH)"], 0)
        show_point_labels = st.toggle("Tampilkan angka di titik grafik", value=False)

        st.markdown("---")
        st.subheader("Filter (multi pilih)")
        with st.form("filter_form", clear_on_submit=False):
            category_sel = st.multiselect("CATEGORY (COUNTRY)", options_for(df_all, "COUNTRY"), default=[])
            transaksi_sel = st.multiselect("TRANSAKSI", options_for(df_all, "TRANSAKSI"), default=[])
            team_sel = st.multiselect("TEAM", options_for(df_all, "TEAM"), default=[])
            product_sel = st.multiselect("PRODUCT", options_for(df_all, "PRODUCT"), default=[])
            brand_sel = st.multiselect("BRAND", options_for(df_all, "BRAND"), default=[])
            platform_sel = st.multiselect("PLATFORM (NAMA CUSTOMER)", options_for(df_all, "PLATFORM"), default=[])
            apply_clicked = st.form_submit_button("✅ Apply Filter")

    if "filters" not in st.session_state:
        st.session_state["filters"] = {"COUNTRY": [], "TRANSAKSI": [], "TEAM": [], "PRODUCT": [], "BRAND": [], "PLATFORM": []}
    if apply_clicked:
        st.session_state["filters"] = {
            "COUNTRY": category_sel,
            "TRANSAKSI": transaksi_sel,
            "TEAM": team_sel,
            "PRODUCT": product_sel,
            "BRAND": brand_sel,
            "PLATFORM": platform_sel,
        }

    flt = st.session_state["filters"]


    def apply_all_filters(df: pd.DataFrame) -> pd.DataFrame:
        out = df
        out = apply_multifilter(out, "COUNTRY", flt["COUNTRY"])
        out = apply_multifilter(out, "TRANSAKSI", flt["TRANSAKSI"])
        out = apply_multifilter(out, "TEAM", flt["TEAM"])
        out = apply_multifilter(out, "PRODUCT", flt["PRODUCT"])
        out = apply_multifilter(out, "BRAND", flt["BRAND"])
        out = apply_multifilter(out, "PLATFORM", flt["PLATFORM"])
        return out


    df_all_f = apply_all_filters(df_all)
    df_a_f = apply_all_filters(df_a)
    df_b_f = apply_all_filters(df_b)

    df_last, df_this, label_last, label_this = build_period_frames(df_all_f, compare_mode, df_a_f, df_b_f)

    metric_col = "QTY" if metric_choice.startswith("Qty") else "JUMLAH"
    metric_name = "Qty" if metric_col == "QTY" else "Sales (IDR)"

    # =========================
    # KPI
    # =========================
    k_last = compute_kpis_cached(df_last)
    k_this = compute_kpis_cached(df_this)

    sales_g = safe_growth_pct(k_this["sales"], k_last["sales"])
    orders_g = safe_growth_pct(k_this["orders"], k_last["orders"])
    qty_g = safe_growth_pct(k_this["qty"], k_last["qty"])
    aov_g = safe_growth_pct(k_this["aov"], k_last["aov"]) if (not pd.isna(k_this["aov"]) and not pd.isna(k_last["aov"])) else None

    st.subheader("Ringkasan Periode")
    c1, c2, c3, c4, c5, c6 = st.columns(6)


    def summary_card(title: str, value: str):
        st.markdown(
            f"""
    <div class="card">
      <div class="card-title">{title}</div>
      <div class="card-value" style="font-size:15px">{value}</div>
    </div>
    """,
            unsafe_allow_html=True,
        )


    with c1:
        summary_card(label_last, f"{df_last['TGL'].min()} → {df_last['TGL'].max()}" if len(df_last) else "-")
    with c2:
        summary_card(label_this, f"{df_this['TGL'].min()} → {df_this['TGL'].max()}" if len(df_this) else "-")
    with c3:
        summary_card("Rows Periode Lalu", f"{len(df_last):,}".replace(",", "."))
    with c4:
        summary_card("Rows Periode Ini", f"{len(df_this):,}".replace(",", "."))

    st.markdown("<hr/>", unsafe_allow_html=True)

    st.subheader("KPI Utama")
    kpi_html = f"""
    <div class="kpi-grid">
      <div class="card">
        <div class="card-title">Total Sales (Periode Ini)</div>
        <div class="card-value">{format_idr(k_this["sales"])}</div>
        <div class="card-sub {kpi_delta_class(sales_g)}">{growth_label(sales_g)}</div>
      </div>
      <div class="card">
        <div class="card-title">Orders</div>
        <div class="card-value">{format_int_id(k_this["orders"])}</div>
        <div class="card-sub {kpi_delta_class(orders_g)}">{growth_label(orders_g)}</div>
      </div>
      <div class="card">
        <div class="card-title">Total Qty</div>
        <div class="card-value">{format_int_id(k_this["qty"])}</div>
        <div class="card-sub {kpi_delta_class(qty_g)}">{growth_label(qty_g)}</div>
      </div>
      <div class="card">
        <div class="card-title">AOV</div>
        <div class="card-value">{format_idr(k_this["aov"])}</div>
        <div class="card-sub {kpi_delta_class(aov_g)}">{growth_label(aov_g)}</div>
      </div>
      <div class="card">
        <div class="card-title">Retur (lines)</div>
        <div class="card-value">{format_int_id(k_this["returns"])}</div>
      </div>
    </div>
    """.replace(",", ".")
    st.markdown(kpi_html, unsafe_allow_html=True)

    # =========================
    # NEW TABLE #1: % TEAM TURUN
    # =========================
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader("📉 Ringkasan Pergerakan TEAM (QTY)")
    down_tbl = team_down_ratio_table_cached(df_last, df_this)
    st.dataframe(
        down_tbl,
        use_container_width=True,
        hide_index=True,
        column_config={"% TEAM Turun": st.column_config.NumberColumn(format="%.2f")},
    )

    st.markdown("<hr/>", unsafe_allow_html=True)

    COLOR_MAP_PERIOD = {
        "Bulan Ini": "#1f77b4",
        "Bulan Lalu": "#aec7e8",
    }

    # =========================
    # Trend chart (Day-of-Month comparison)
    # =========================
    st.subheader(f"Tren Harian ({metric_name})")


    def day_of_month_series(df: pd.DataFrame, label: str) -> pd.DataFrame:
        tmp = df.copy()
        tmp["DAY"] = pd.to_datetime(tmp["TGL"]).dt.day
        g = tmp.groupby("DAY", as_index=False).agg(VALUE=(metric_col, "sum"))
        g["PERIODE"] = label
        return g


    all_days = pd.DataFrame({"DAY": list(range(1, 32))})
    trend_dom = pd.concat(
        [
            all_days.merge(day_of_month_series(df_last, label_last), on="DAY", how="left").assign(PERIODE=label_last),
            all_days.merge(day_of_month_series(df_this, label_this), on="DAY", how="left").assign(PERIODE=label_this),
        ],
        ignore_index=True,
    )
    trend_dom["VALUE"] = trend_dom["VALUE"].fillna(0.0)

    if show_point_labels:
        trend_dom["LABEL"] = trend_dom["VALUE"].apply(compact_number)
        fig = px.line(
            trend_dom,
            x="DAY",
            y="VALUE",
            color="PERIODE",
            markers=True,
            text="LABEL",
            color_discrete_map=COLOR_MAP_PERIOD,
        )
        fig.update_traces(textposition="top center")
    else:
        fig = px.line(
            trend_dom,
            x="DAY",
            y="VALUE",
            color="PERIODE",
            markers=True,
            color_discrete_map=COLOR_MAP_PERIOD,
        )

    fig.update_layout(
        xaxis_title="Tanggal (Day of Month)",
        yaxis_title=metric_name,
        legend_title_text="",
        xaxis=dict(dtick=1),
    )
    st.plotly_chart(fig, use_container_width=True)

    # =========================
    # Cumulative chart
    # =========================
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader(f"Statistik Kumulatif ({metric_name})")

    trend_cum = trend_dom.copy()
    trend_cum = trend_cum.sort_values(["PERIODE", "DAY"]).copy()
    trend_cum["CUM_VALUE"] = trend_cum.groupby("PERIODE")["VALUE"].cumsum()

    # super clean compare hover: tampilkan Bulan Ini + Bulan Lalu + Delta dalam satu hover
    trend_compare = (
        trend_cum.pivot(index="DAY", columns="PERIODE", values="CUM_VALUE")
        .reset_index()
        .rename_axis(None, axis=1)
    )

    if "Bulan Ini" not in trend_compare.columns:
        trend_compare["Bulan Ini"] = 0
    if "Bulan Lalu" not in trend_compare.columns:
        trend_compare["Bulan Lalu"] = 0

    trend_compare["Bulan Ini"] = trend_compare["Bulan Ini"].fillna(0)
    trend_compare["Bulan Lalu"] = trend_compare["Bulan Lalu"].fillna(0)
    trend_compare["DELTA"] = trend_compare["Bulan Ini"] - trend_compare["Bulan Lalu"]

    trend_cum = trend_cum.merge(
        trend_compare[["DAY", "Bulan Ini", "Bulan Lalu", "DELTA"]],
        on="DAY",
        how="left"
    )

    trend_cum["BULAN_INI_TXT"] = trend_cum["Bulan Ini"].apply(compact_number)
    trend_cum["BULAN_LALU_TXT"] = trend_cum["Bulan Lalu"].apply(compact_number)
    trend_cum["DELTA_TXT"] = trend_cum["DELTA"].apply(
        lambda x: f"+{compact_number(x)}" if x > 0 else compact_number(x)
    )

    fig_cum = px.line(
        trend_cum,
        x="DAY",
        y="CUM_VALUE",
        color="PERIODE",
        markers=True,
        color_discrete_map=COLOR_MAP_PERIOD,
        custom_data=["BULAN_INI_TXT", "BULAN_LALU_TXT", "DELTA_TXT"],
    )

    fig_cum.update_traces(
        hovertemplate=(
            "<b>Hari %{x}</b><br>"
            "Bulan Ini: %{customdata[0]}<br>"
            "Bulan Lalu: %{customdata[1]}<br>"
            "Delta: %{customdata[2]}<extra>%{fullData.name}</extra>"
        )
    )

    fig_cum.update_layout(
        xaxis_title="Tanggal (Day of Month)",
        yaxis_title=f"Kumulatif {metric_name}",
        legend_title_text="",
        xaxis=dict(dtick=1),
    )
    st.plotly_chart(fig_cum, use_container_width=True)


    # =========================
    # Pareto + Delta + Comparison
    # =========================
    st.markdown("<hr/>", unsafe_allow_html=True)

    pareto_dim = st.selectbox(
        "Filter Pareto",
        ["PLATFORM", "TEAM", "PRODUCT", "COUNTRY", "BRAND", "TRANSAKSI", "AREA"],
        index=1,
    )

    pareto_top_n = 30

    def build_pareto_comparison(df_this: pd.DataFrame, df_last: pd.DataFrame, dim_col: str, value_col: str, top_n: int = 10):
        this_agg = (
            df_this.groupby(dim_col, as_index=False)
            .agg(THIS_VALUE=(value_col, "sum"))
        )
        last_agg = (
            df_last.groupby(dim_col, as_index=False)
            .agg(LAST_VALUE=(value_col, "sum"))
        )

        comp = this_agg.merge(last_agg, on=dim_col, how="outer").fillna(0.0)
        comp[dim_col] = comp[dim_col].astype(str).str.strip()
        comp = comp[comp[dim_col].ne("")].copy()
        comp = comp.sort_values("THIS_VALUE", ascending=False).head(top_n).copy()

        if comp.empty:
            return comp

        total_this = comp["THIS_VALUE"].sum()
        total_last = comp["LAST_VALUE"].sum()

        comp["THIS_SHARE"] = np.where(total_this != 0, comp["THIS_VALUE"] / total_this * 100.0, 0.0)
        comp["LAST_SHARE"] = np.where(total_last != 0, comp["LAST_VALUE"] / total_last * 100.0, 0.0)
        comp["PARETO_THIS"] = comp["THIS_SHARE"].cumsum()
        comp["PARETO_LAST"] = comp["LAST_SHARE"].cumsum()
        comp["CUM_STORE_COUNT"] = np.arange(1, len(comp) + 1)
        comp["DELTA_SHARE"] = comp["THIS_SHARE"] - comp["LAST_SHARE"]
        comp["DELTA_LABEL"] = comp["DELTA_SHARE"].map(lambda x: f"{x:+.1f}%")

        comp["BAR_HOVER"] = comp.apply(
            lambda r: (
                f"<b>{r[dim_col]}</b><br>"
                f"{label_this}: {compact_number(r['THIS_VALUE'])}<br>"
                f"{label_last}: {compact_number(r['LAST_VALUE'])}<br>"
                f"Kontribusi {label_this}: {r['THIS_SHARE']:.2f}%<br>"
                f"Kontribusi {label_last}: {r['LAST_SHARE']:.2f}%<br>"
                f"Delta kontribusi: {r['DELTA_SHARE']:+.2f}%"
            ),
            axis=1,
        )

        return comp

    pareto_df = build_pareto_comparison(df_this, df_last, pareto_dim, metric_col, pareto_top_n)

    if pareto_df.empty:
        st.info("Belum ada data untuk grafik Pareto pada filter saat ini.")
    else:

        fig_pareto = make_subplots(specs=[[{"secondary_y": True}]])

        fig_pareto.add_trace(
            go.Bar(
                x=pareto_df[pareto_dim],
                y=pareto_df["THIS_VALUE"],
                name=label_this,
                marker_color="#8ecae6",
                text=pareto_df["DELTA_LABEL"],
                textposition="outside",
                hovertext=pareto_df["BAR_HOVER"],
                hovertemplate="%{hovertext}<extra></extra>",
            ),
            secondary_y=False,
        )

        fig_pareto.add_trace(
            go.Scatter(
                x=pareto_df[pareto_dim],
                y=pareto_df["PARETO_THIS"],
                name=f"Pareto {label_this}",
                mode="lines+markers",
                line=dict(color="#1f77b4", width=3),
                marker=dict(symbol="circle", size=7, color="#1f77b4"),
                customdata=pareto_df[["CUM_STORE_COUNT"]],
                hovertemplate=(
                    "jumlah toko: %{customdata[0]}<br>"
                    f"Pareto {label_this}: %{{y:.2f}}%<extra></extra>"
                ),
            ),
            secondary_y=True,
        )

        fig_pareto.add_trace(
            go.Scatter(
                x=pareto_df[pareto_dim],
                y=pareto_df["PARETO_LAST"],
                name=f"Pareto {label_last}",
                mode="lines+markers",
                line=dict(color="#f59e0b", width=2.5, dash="dash"),
                marker=dict(symbol="x", size=8, color="#f59e0b"),
                customdata=pareto_df[["CUM_STORE_COUNT"]],
                hovertemplate=(
                    "jumlah toko: %{customdata[0]}<br>"
                    f"Pareto {label_last}: %{{y:.2f}}%<extra></extra>"
                ),
            ),
            secondary_y=True,
        )

        fig_pareto.add_hline(
            y=80,
            line_width=1.5,
            line_dash="solid",
            line_color="#3b82f6",
            opacity=0.85,
            secondary_y=True,
        )

        fig_pareto.update_layout(
            title="Pareto + Delta + Comparison (This Vs Last Month)",
            hovermode="x unified",
            legend_title_text="",
            xaxis_title=pareto_dim.title(),
            yaxis_title=f"{metric_name} ({label_this})",
            margin=dict(l=40, r=40, t=70, b=40),
        )

        fig_pareto.update_yaxes(
            title_text=f"{metric_name} ({label_this})",
            secondary_y=False,
            showgrid=True,
            gridcolor="rgba(0,0,0,0.08)",
        )
        fig_pareto.update_yaxes(
            title_text="Cumulative (%)",
            range=[0, 105],
            ticksuffix="%",
            secondary_y=True,
            showgrid=False,
        )

        st.plotly_chart(fig_pareto, use_container_width=True)


    st.markdown("<hr/>", unsafe_allow_html=True)

    # =========================
    # NEW TABLE #2: TEAM DRIVER ANALYSIS
    # =========================
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader("🧠 Analisis Penyebab Perubahan")
    st.caption("Untuk TEAM yang TURUN: ditampilkan top driver yang paling narik turun. Untuk TEAM yang NAIK: top driver yang paling narik naik.")

    topk = st.slider("Top driver per kategori", 1, 10, 3, 1)
    analysis_df = team_driver_analysis_table_cached(df_last, df_this, top_k=topk)

    st.dataframe(
        style_growth_pct_df(analysis_df),
        use_container_width=True,
        height=520,
    )

    # =========================
    # Top tables
    # =========================
    c1, c2 = st.columns(2)
    with c1:
        small_title(f"Top {top_n} TEAM", f"(by {metric_col})")
        render_html_table(top_table_cached(df_this, df_last, "TEAM", metric_col, top_n))
    with c2:
        small_title(f"Top {top_n} PRODUCT", f"(by {metric_col})")
        render_html_table(top_table_cached(df_this, df_last, "PRODUCT", metric_col, top_n))

    c3, c4 = st.columns(2)
    with c3:
        small_title(f"Top {top_n} BRAND", f"(by {metric_col})")
        render_html_table(top_table_cached(df_this, df_last, "BRAND", metric_col, top_n))
    with c4:
        small_title(f"Top {top_n} TRANSAKSI", f"(by {metric_col})")
        render_html_table(top_table_cached(df_this, df_last, "TRANSAKSI", metric_col, top_n))

    c5, c6 = st.columns(2)
    with c5:
        small_title(f"Top {top_n} SKU", "(SPESIFIKASI)")
        render_html_table(top_table_cached(df_this, df_last, "SPESIFIKASI", metric_col, top_n), table_class="sku-table")
    with c6:
        small_title(f"Top {top_n} PLATFORM", "(sumber: NAMA CUSTOMER)")
        render_html_table(top_table_cached(df_this, df_last, "PLATFORM", metric_col, top_n))

    st.markdown("<hr/>", unsafe_allow_html=True)

    # =========================
    # TEAM PERFORMANCE (3 columns)
    # =========================
    st.subheader("📊 Team Performance Insight (QTY)")

    team_last = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
    team_this = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
    team = team_last.merge(team_this, on="TEAM", how="outer").fillna(0.0)
    team["DELTA_QTY"] = team["QTY_INI"] - team["QTY_LALU"]
    team["GROWTH_PCT"] = team.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

    under = team[team["QTY_INI"] < 30].copy().sort_values(["QTY_INI", "GROWTH_PCT"], ascending=[True, True])

    oto_team = (
        df_this.groupby("TEAM", as_index=False)
        .agg(
            OTO_YES_LINES=("OTO_YES", "sum"),
            TOTAL_LINES=("OTO_YES", "count"),
            QTY_INI=("QTY", "sum"),
        )
    )
    oto_team["OTO_RATE"] = np.where(oto_team["TOTAL_LINES"] > 0, oto_team["OTO_YES_LINES"] / oto_team["TOTAL_LINES"] * 100.0, 0.0)
    oto_team = oto_team.sort_values(["OTO_YES_LINES", "OTO_RATE"], ascending=[False, False])

    top_all = team.copy().sort_values(["GROWTH_PCT", "QTY_INI"], ascending=[False, False])


    def prep_team_view(df_in: pd.DataFrame) -> pd.DataFrame:
        d = df_in.copy()
        d["QTY Lalu"] = d["QTY_LALU"].map(format_int_id)
        d["QTY Ini"] = d["QTY_INI"].map(format_int_id)
        d["Delta"] = d["DELTA_QTY"].map(format_int_id)
        d["Growth %"] = d["GROWTH_PCT"].apply(lambda x: float(x) if (x is not None and not pd.isna(x)) else np.nan)
        return d[["TEAM", "QTY Lalu", "QTY Ini", "Delta", "Growth %"]]


    def prep_oto_view(df_in: pd.DataFrame) -> pd.DataFrame:
        d = df_in.copy()
        d["OTO YES (lines)"] = d["OTO_YES_LINES"].astype(int)
        d["OTO Rate %"] = d["OTO_RATE"]
        d["QTY (Periode Ini)"] = d["QTY_INI"].map(format_int_id)
        return d[["TEAM", "OTO YES (lines)", "OTO Rate %", "QTY (Periode Ini)"]]


    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("### 🏆 Top Performer (All TEAM)")
        st.caption("Scroll & klik header kolom untuk sort (Growth% / QTY).")
        df_top = prep_team_view(top_all).copy()
        st.dataframe(style_growth_pct_df(df_top), use_container_width=True, height=420)

    with col2:
        st.markdown("### ⚠️ Under Perform (QTY < 30)")
        st.caption("Team dengan QTY periode ini di bawah 30.")
        df_under = prep_team_view(under).copy()
        st.dataframe(style_growth_pct_df(df_under), use_container_width=True, height=420)

    with col3:
        st.markdown('### 🚫 Team sering OTO "YES"')
        st.caption('Urut berdasarkan jumlah OTO == "YES" (periode ini).')
        st.dataframe(
            prep_oto_view(oto_team),
            use_container_width=True,
            height=420,
            column_config={"OTO Rate %": st.column_config.NumberColumn(format="%.2f")},
            hide_index=True,
        )

    # =========================
    # EXTRA INSIGHT (tambahan)
    # =========================
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader("🧠 Insight Tambahan (QTY / Retur / AREA)")

    # ---------- 1) TEAM qty besar tapi turun ----------
    st.markdown("### 1) TEAM QTY besar tapi turun")

    team_qty = (
        df_this.groupby("TEAM", as_index=False)
        .agg(QTY_INI=("QTY", "sum"))
        .merge(
            df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum")),
            on="TEAM",
            how="left",
        )
        .fillna(0.0)
    )
    team_qty["DELTA"] = team_qty["QTY_INI"] - team_qty["QTY_LALU"]
    team_qty["GROWTH_PCT"] = team_qty.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

    # ambil kandidat TEAM dengan QTY ini terbesar, lalu filter yang turun
    TOP_BIG = 30  # bisa kamu ubah 20/50
    big_down = (
        team_qty.sort_values("QTY_INI", ascending=False)
        .head(TOP_BIG)
        .query("DELTA < 0")
        .copy()
        .sort_values("DELTA", ascending=True)
        .head(20)
    )

    if len(big_down) == 0:
        st.info("Tidak ada TEAM 'QTY besar tapi turun' pada filter & periode saat ini.")
    else:
        big_down_view = big_down.copy()
        big_down_view["QTY Ini"] = big_down_view["QTY_INI"].map(format_int_id)
        big_down_view["QTY Lalu"] = big_down_view["QTY_LALU"].map(format_int_id)
        big_down_view["Delta"] = big_down_view["DELTA"].map(format_int_id)
        big_down_view["Growth"] = big_down_view["GROWTH_PCT"].apply(growth_badge_html)
        big_down_view = big_down_view[["TEAM", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
        render_html_table(big_down_view)

    # ---------- 2) TEAM paling banyak retur ----------
    st.markdown("### 2) TEAM paling banyak retur")

    ret_this = df_this[df_this["ROW_TYPE"] == "RETUR"].copy()

    if len(ret_this) == 0:
        st.info("Tidak ada data RETUR pada periode ini (berdasarkan kolom STATUS).")
    else:
        ret_team = (
            ret_this.groupby("TEAM", as_index=False)
            .agg(
                Retur_Lines=("TEAM", "count"),
                Retur_QTY=("QTY", "sum"),
            )
            .sort_values(["Retur_Lines", "Retur_QTY"], ascending=[False, True])
            .head(20)
            .copy()
        )
        # QTY retur biasanya negatif, biar enak lihat pakai ABS
        ret_team["Retur_QTY (abs)"] = ret_team["Retur_QTY"].abs().map(format_int_id)
        ret_team["Retur_Lines"] = ret_team["Retur_Lines"].map(format_int_id)
        ret_team = ret_team[["TEAM", "Retur_Lines", "Retur_QTY (abs)"]]
        render_html_table(ret_team)

    # ---------- 3) Perform AREA (Naik/Turun) ----------
    st.markdown("### 3) Perform AREA (Naik / Turun)")

    if "AREA" not in df_this.columns or "AREA" not in df_last.columns:
        st.warning("Kolom 'AREA' tidak ditemukan di data yang terbaca. (Pastikan Excel punya kolom AREA & tidak terbuang saat cleaning).")
    else:
        area_this = df_this.groupby("AREA", as_index=False).agg(QTY_INI=("QTY", "sum"))
        area_last = df_last.groupby("AREA", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        area = area_this.merge(area_last, on="AREA", how="outer").fillna(0.0)

        area["DELTA"] = area["QTY_INI"] - area["QTY_LALU"]
        area["GROWTH_PCT"] = area.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        # tampilkan top naik & top turun
        top_up = area.sort_values("DELTA", ascending=False).head(10).copy()
        top_dn = area.sort_values("DELTA", ascending=True).head(10).copy()

        colA, colB = st.columns(2)
        with colA:
            st.markdown("#### 🔼 Top AREA Naik (QTY)")
            v = top_up.copy()
            v["QTY Ini"] = v["QTY_INI"].map(format_int_id)
            v["QTY Lalu"] = v["QTY_LALU"].map(format_int_id)
            v["Delta"] = v["DELTA"].map(format_int_id)
            v["Growth"] = v["GROWTH_PCT"].apply(growth_badge_html)
            v = v[["AREA", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
            render_html_table(v)

        with colB:
            st.markdown("#### 🔽 Top AREA Turun (QTY)")
            v = top_dn.copy()
            v["QTY Ini"] = v["QTY_INI"].map(format_int_id)
            v["QTY Lalu"] = v["QTY_LALU"].map(format_int_id)
            v["Delta"] = v["DELTA"].map(format_int_id)
            v["Growth"] = v["GROWTH_PCT"].apply(growth_badge_html)
            v = v[["AREA", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
            render_html_table(v)


# ============================================================
# EXTERNAL MODULE: render_analisa_produk_stok
# ============================================================
def render_analisa_produk_stok():
    PERIODS = ["7DAY", "14DAY", "30DAY"]
    DIVISIONS = ["DIV03", "DIV04", "DIV05"]

    MPLSSR_DIV_COLS = {
        "7DAY": {"DIV03": "03 OLP", "DIV04": "04 MOD", "DIV05": "05 OLR"},
        "14DAY": {"DIV03": "03 OLP.1", "DIV04": "04 MOD.1", "DIV05": "05 OLR.1"},
        "30DAY": {"DIV03": "03 OLP.2", "DIV04": "04 MOD.2", "DIV05": "05 OLR.2"},
    }

    VALID_PRICELIST_SHEETS = [
        "LAPTOP",
        "TELCO",
        "PC HOM ELE",
        "SOF COM SUP",
        "ACC",
        "SER OTH CON",
    ]

    PRICE_SEGMENTS = [
        (0, 1_000_000, "< 1 JUTA"),
        (1_000_000, 1_500_000, "1 - 1.5 JUTA"),
        (1_500_000, 2_000_000, "1.5 - 2 JUTA"),
        (2_000_000, 2_500_000, "2 - 2.5 JUTA"),
        (2_500_000, 3_000_000, "2.5 - 3 JUTA"),
        (3_000_000, 4_000_000, "3 - 4 JUTA"),
        (4_000_000, 5_000_000, "4 - 5 JUTA"),
        (5_000_000, 7_000_000, "5 - 7 JUTA"),
        (7_000_000, 10_000_000, "7 - 10 JUTA"),
        (10_000_000, 12_500_000, "10 - 12.5 JUTA"),
        (12_500_000, 15_000_000, "12.5 - 15 JUTA"),
        (15_000_000, 20_000_000, "15 - 20 JUTA"),
        (20_000_000, 25_000_000, "20 - 25 JUTA"),
        (25_000_000, 30_000_000, "25 - 30 JUTA"),
        (30_000_000, 40_000_000, "30 - 40 JUTA"),
        (40_000_000, np.inf, "40 JUTA - UP"),
    ]

    SEGMENT_ORDER = [label for _, _, label in PRICE_SEGMENTS]

    st.markdown(
        """
        <style>
        .block-container {padding-top: 1rem; padding-bottom: 1rem;}
        .upload-card-wrap {
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            background: #ffffff;
            padding: 14px;
            margin-bottom: 16px;
        }
        .table-card {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            padding: 10px;
            margin-bottom: 14px;
        }
        .alert-wrap {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            overflow: auto;
            max-height: 520px;
        }
        table.alert-table {
            border-collapse: collapse;
            width: max-content;
            min-width: 100%;
            table-layout: fixed;
            font-size: 12px;
        }
        table.alert-table th, table.alert-table td {
            border: 1px solid #e5e7eb;
            padding: 6px 8px;
            text-align: left;
            white-space: nowrap;
        }
        table.alert-table thead th {
            position: sticky;
            top: 0;
            background: #f8fafc;
            z-index: 2;
        }
        .main-fixed-wrap {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            overflow: auto;
            max-height: 520px;
        }
        table.main-fixed {
            border-collapse: collapse;
            width: max-content;
            min-width: 100%;
            table-layout: fixed;
            font-size: 12px;
        }
        table.main-fixed th, table.main-fixed td {
            border: 1px solid #e5e7eb;
            padding: 6px 8px;
            text-align: left;
            white-space: nowrap;
        }
        table.main-fixed thead th {
            position: sticky;
            top: 0;
            background: #f8fafc;
            z-index: 2;
        }
        table.main-fixed th:nth-child(1),
        table.main-fixed td:nth-child(1) { min-width: 180px; }
        table.main-fixed th:nth-child(2),
        table.main-fixed td:nth-child(2) { min-width: 90px; }
        table.main-fixed th:nth-child(3),
        table.main-fixed td:nth-child(3) { min-width: 80px; }
        table.main-fixed th:nth-child(4),
        table.main-fixed td:nth-child(4) {
            min-width: 420px;
            max-width: 420px;
            white-space: normal;
            word-break: break-word;
            line-height: 1.25;
        }
        table.main-fixed th:nth-child(5),
        table.main-fixed td:nth-child(5) { min-width: 80px; }
        table.main-fixed th:nth-child(6),
        table.main-fixed td:nth-child(6),
        table.main-fixed th:nth-child(7),
        table.main-fixed td:nth-child(7),
        table.main-fixed th:nth-child(8),
        table.main-fixed td:nth-child(8),
        table.main-fixed th:nth-child(9),
        table.main-fixed td:nth-child(9) { min-width: 70px; }
        .bg-red {
            background: #ffebee;
            color: #c62828;
            font-weight: 700;
        }
        .status-refill {
            background: #ffebee;
            color: #c62828;
            font-weight: 700;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    def normalize_text(series: pd.Series) -> pd.Series:
        return (
            series.astype(str)
            .str.strip()
            .str.upper()
            .replace({"NAN": np.nan, "NONE": np.nan, "": np.nan})
        )

    def to_num(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce")

    def format_thousands_id(value):
        try:
            num = pd.to_numeric(value, errors="coerce")
            if pd.isna(num):
                return ""
            return f"{int(round(float(num) / 1000)):,}".replace(",", ".")
        except Exception:
            return ""

    def format_units_id(value):
        try:
            num = pd.to_numeric(value, errors="coerce")
            if pd.isna(num):
                return ""
            return f"{int(round(float(num))):,}".replace(",", ".")
        except Exception:
            return ""

    def price_segment(price: float) -> str:
        if pd.isna(price):
            return "UNKNOWN"
        for low, high, label in PRICE_SEGMENTS:
            if low <= float(price) < high:
                return label
        return "UNKNOWN"

    def segment_sort_key(label: str) -> int:
        try:
            return SEGMENT_ORDER.index(label)
        except ValueError:
            return len(SEGMENT_ORDER)

    def first_row_contains_text(df: pd.DataFrame, text: str):
        target = str(text).strip().upper()
        for idx in range(len(df)):
            row_text = df.iloc[idx].astype(str).str.upper()
            if row_text.str.contains(target, na=False).any():
                return idx
        return None

    def _ffill_header(values: List) -> List:
        out = []
        last = None
        for v in values:
            if pd.notna(v) and str(v).strip() != "":
                last = str(v).strip().upper()
                out.append(last)
            else:
                out.append(last)
        return out


    def _norm_header_cell(value) -> str:
        if pd.isna(value) or str(value).strip() == "":
            return ""
        return str(value).strip().upper()

    def find_div05_stock_columns(columns: List[str], excel_row3_raw: List, excel_row4_raw: List) -> List[str]:
        # Header bisa merge cell, jadi harus di-forward fill ke kanan.
        row3_vals = [_norm_header_cell(v) for v in _ffill_header(excel_row3_raw)]
        row4_vals = [_norm_header_cell(v).replace(" ", "") for v in _ffill_header(excel_row4_raw)]

        # 05 OLR dimulai dari area RAM di row 3 dan dihitung sampai kolom paling kanan.
        ram_start = next((i for i, v in enumerate(row3_vals) if "RAM" in v), None)

        # Fallback: kalau RAM tidak ditemukan, mulai dari kolom pertama 5B di row 4.
        if ram_start is None:
            ram_start = next((i for i, v in enumerate(row4_vals) if v == "5B"), None)

        if ram_start is None:
            return []

        return [columns[i] for i in range(ram_start, len(columns)) if 0 <= i < len(columns)]

    def area_code_matches(value, prefixes: List[str]) -> bool:
        if pd.isna(value):
            return False
        txt = str(value).strip().upper().replace(" ", "")
        return any(txt.startswith(p) for p in prefixes)

    def normalize_warehouse_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        txt = re.sub(r"0+(?=\d)", "", txt)
        txt = txt.replace("0A", "A").replace("0B", "B").replace("0C", "C")
        return txt


    def normalize_sales_pivot_gudang(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+[A-Z]?)$", txt)
        if m:
            prefix = m.group(1)
            suffix = m.group(2)
            return f"{prefix} {suffix}"
        return txt


    def normalize_display_team_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace("_", "").replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+)([A-Z])?$", txt)
        if m:
            prefix = m.group(1)
            number = int(m.group(2))
            suffix = m.group(3) if m.group(3) else "A"
            return f"{prefix} {number}{suffix}"
        return txt


    def normalize_team_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+)$", txt)
        if m:
            return f"{m.group(1)} {int(m.group(2))}A"
        m = re.match(r"([A-Z]+)0*(\d+)([A-Z])$", txt)
        if m:
            return f"{m.group(1)} {int(m.group(2))}{m.group(3)}"
        return txt


    def normalize_combined_warehouse_code(group, wh=None) -> str:
        parts = []
        for val in [group, wh]:
            if pd.isna(val) or str(val).strip() == "":
                continue
            txt = str(val).strip().upper()
            txt = txt.replace(" ", "")
            parts.append(txt)
        return "".join(parts)

    def normalize_team_lookup_key(value) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip().upper().replace(" ", "").replace("-", "")

    def ensure_datetime(series: pd.Series) -> pd.Series:
        return pd.to_datetime(series, errors="coerce", dayfirst=True)


    def get_period_date_range(df: pd.DataFrame, period: str):
        if df is None or df.empty or "TGL" not in df.columns:
            return pd.NaT, pd.NaT

        tgl = pd.to_datetime(df["TGL"], errors="coerce").dropna()
        if tgl.empty:
            return pd.NaT, pd.NaT

        max_date = tgl.max().normalize()
        days_map = {"7DAY": 7, "14DAY": 14, "30DAY": 30}
        days = days_map.get(period, 7)
        start_date = max_date - pd.Timedelta(days=days - 1)
        return start_date, max_date


    # =========================================================
    # MPLSSR
    # =========================================================
    def load_mplssr(file) -> pd.DataFrame:
        df = pd.read_excel(file, sheet_name="ALL", header=1)
        df = df.iloc[4:].copy().reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.duplicated()]

        base_cols = ["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI"]
        for c in base_cols:
            if c not in df.columns:
                df[c] = np.nan

        div_cols = []
        for p in PERIODS:
            for col in MPLSSR_DIV_COLS[p].values():
                if col in df.columns:
                    div_cols.append(col)

        df = df[base_cols + div_cols].copy()

        for c in base_cols:
            df[c] = normalize_text(df[c])

        df = df[df["KODE BARANG"].notna()].copy()
        df = df[~df["KODE BARANG"].isin(["TOTAL", "SHARE%"])]

        rows = []
        for period in PERIODS:
            for div, col in MPLSSR_DIV_COLS[period].items():
                if col not in df.columns:
                    continue
                tmp = df[["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI"]].copy()
                tmp["PERIOD"] = period
                tmp["DIVISION"] = div
                tmp["QTY"] = to_num(df[col]).fillna(0)
                rows.append(tmp)

        out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(
            columns=["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI", "PERIOD", "DIVISION", "QTY"]
        )
        out["MERGE_KEY"] = normalize_text(out["KODE BARANG"])
        out["SKU NO"] = out["MERGE_KEY"]
        return out

    # =========================================================
    # PRICELIST
    # =========================================================

    def parse_pricelist_sheet(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
        raw = xls.parse(sheet_name=sheet_name, header=None).copy()

        if sheet_name.upper() == "LAPTOP":
            coming_idx = first_row_contains_text(raw, "COMING")
            end_coming_idx = first_row_contains_text(raw, "END COMING")
            if coming_idx is not None and end_coming_idx is not None and end_coming_idx >= coming_idx:
                raw = raw.drop(index=range(coming_idx, end_coming_idx + 1)).reset_index(drop=True)

        row1 = raw.iloc[1].tolist()
        row2_raw = raw.iloc[2].tolist()
        row2 = _ffill_header(row2_raw)
        row3_raw = raw.iloc[3].tolist()
        row3 = [str(x).strip().upper() if pd.notna(x) and str(x).strip() != "" else None for x in row3_raw]

        columns = []
        for i, v in enumerate(row1):
            v1 = str(v).strip().upper() if pd.notna(v) and str(v).strip() != "" else None
            if v1 is not None:
                columns.append(v1)
            elif row2[i] is not None and row3[i] is not None:
                columns.append(f"{row2[i]}__{row3[i]}")
            elif row3[i] is not None:
                columns.append(row3[i])
            else:
                columns.append(f"COL_{i}")

        df = raw.iloc[5:].copy().reset_index(drop=True)
        df.columns = columns

        for c in ["SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "TOT", "M3"]:
            if c not in df.columns:
                df[c] = np.nan

        df["SKU NO"] = normalize_text(df["SKU NO"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODEBARANG"] = normalize_text(df["KODEBARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])

        df = df[df["KODEBARANG"].notna()].copy()
        df = df[~df["KODEBARANG"].isin(["TOTAL"])]

        stock03_cols = [columns[i] for i, area in enumerate(row3) if area_code_matches(area, ["3", "03"])]
        stock04_cols = [columns[i] for i, area in enumerate(row3) if area_code_matches(area, ["4", "04"])]
        stock05_cols = find_div05_stock_columns(columns, row2_raw, row3_raw)

        df["PRICE"] = to_num(df["M3"]) * 1000
        df["STOK_DIV03"] = df[stock03_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock03_cols else 0
        df["STOK_DIV04"] = df[stock04_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock04_cols else 0
        df["STOK_DIV05"] = df[stock05_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock05_cols else 0
        df["CATEGORY"] = sheet_name.upper()
        df["PRICE_SEGMENT"] = df["PRICE"].apply(price_segment)
        df["MERGE_KEY"] = df["KODEBARANG"]

        return df[[
            "SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "PRICE",
            "STOK_DIV03", "STOK_DIV04", "STOK_DIV05",
            "CATEGORY", "PRICE_SEGMENT", "MERGE_KEY"
        ]]

    def load_pricelist(file) -> pd.DataFrame:
        xls = pd.ExcelFile(file)
        sheets = [s for s in xls.sheet_names if s.upper() in VALID_PRICELIST_SHEETS]
        frames = [parse_pricelist_sheet(xls, s) for s in sheets]
        if not frames:
            return pd.DataFrame(columns=[
                "SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "PRICE",
                "STOK_DIV03", "STOK_DIV04", "STOK_DIV05",
                "CATEGORY", "PRICE_SEGMENT", "MERGE_KEY"
            ])
        out = pd.concat(frames, ignore_index=True)
        out = out.loc[:, ~out.columns.duplicated()]
        out = out.drop_duplicates(subset=["MERGE_KEY"], keep="first")
        return out

    # =========================================================
    # PRICELIST WITH WAREHOUSES
    # =========================================================



    def parse_pricelist_sheet_with_warehouses(xls: pd.ExcelFile, sheet_name: str):
        raw = xls.parse(sheet_name=sheet_name, header=None).copy()

        if sheet_name.upper() == "LAPTOP":
            col_c = raw.iloc[:, 2].astype(str).str.upper().str.strip() if raw.shape[1] > 2 else pd.Series(dtype=str)
            coming_candidates = col_c[col_c.str.contains("COMING", na=False)]
            end_coming_candidates = col_c[col_c.str.contains("END COMING", na=False)]
            if not coming_candidates.empty and not end_coming_candidates.empty:
                coming_idx = int(coming_candidates.index.min())
                end_coming_idx = int(end_coming_candidates.index.max())
                if end_coming_idx >= coming_idx:
                    raw = raw.drop(index=range(coming_idx, end_coming_idx + 1)).reset_index(drop=True)

        # Struktur pricelist:
        # - Row 2: header utama
        # - Row 3: kode gudang utama (mis. RIF, OA)
        # - Row 4: sub-kode gudang (mis. 1A, 2A)
        row1 = raw.iloc[1].tolist()
        row2 = _ffill_header(raw.iloc[2].tolist())
        row3 = _ffill_header(raw.iloc[2].tolist())
        row4 = _ffill_header(raw.iloc[3].tolist()) if len(raw) > 3 else [None] * len(row1)

        columns = []
        warehouse_meta = []

        for i, v in enumerate(row1):
            v1 = str(v).strip().upper() if pd.notna(v) and str(v).strip() != "" else None
            group = str(row3[i]).strip().upper() if pd.notna(row3[i]) and str(row3[i]).strip() != "" else None
            wh = str(row4[i]).strip().upper() if pd.notna(row4[i]) and str(row4[i]).strip() != "" else None

            if v1 is not None:
                columns.append(v1)
                warehouse_meta.append((None, None, None, None))
                continue

            group_clean = _norm_header_cell(group)
            wh_clean = _norm_header_cell(wh).replace(" ", "")

            combined_key = normalize_team_lookup_key(f"{group_clean} {wh_clean}") if group_clean and wh_clean else ""
            combined_label = f"{group_clean} {wh_clean}".strip() if group_clean and wh_clean else (group_clean or "")

            if combined_key:
                columns.append(f"{combined_key}__{i}")
                warehouse_meta.append((group_clean, wh_clean, combined_label, combined_key))
            elif group_clean:
                columns.append(f"{group_clean}__{i}")
                warehouse_meta.append((group_clean, None, group_clean, normalize_team_lookup_key(group_clean)))
            else:
                columns.append(f"COL_{i}")
                warehouse_meta.append((None, None, None, None))

        df = raw.iloc[4:].copy().reset_index(drop=True)
        df.columns = columns

        for c in ["SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "M3"]:
            if c not in df.columns:
                df[c] = np.nan

        df["SKU NO"] = normalize_text(df["SKU NO"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODEBARANG"] = normalize_text(df["KODEBARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])

        df = df[df["KODEBARANG"].notna()].copy()
        df = df[~df["KODEBARANG"].isin(["TOTAL"])].copy()
        df["PRICE"] = to_num(df["M3"]) * 1000

        warehouse_stock_cols = {}
        default_stock_cols = []
        jkt_stock_cols = []

        for i, col in enumerate(columns):
            group, wh, combined_label, combined_key = warehouse_meta[i]
            if group is None:
                continue

            group_key = normalize_team_lookup_key(group)

            if group_key == "DEFAULT":
                default_stock_cols.append(col)
            if group_key == "JKT":
                jkt_stock_cols.append(col)
            if combined_key:
                warehouse_stock_cols.setdefault(combined_key, []).append(col)

        if "BRAND" not in df.columns:
            df["BRAND"] = np.nan
        df["BRAND"] = normalize_text(df["BRAND"])
        df["PRICE_SEGMENT"] = df["PRICE"].apply(price_segment)

        mapped_cols = []
        for cols in warehouse_stock_cols.values():
            mapped_cols.extend(cols)

        keep_cols = [
            "SKU NO", "PRODUCT", "BRAND", "KODEBARANG", "SPESIFIKASI", "PRICE", "PRICE_SEGMENT"
        ] + list(dict.fromkeys(default_stock_cols + jkt_stock_cols + mapped_cols))

        out = df[keep_cols].copy()
        out = out.loc[:, ~out.columns.duplicated()].copy()
        out["DEFAULT_STOCK_TOTAL"] = (
            df.loc[:, ~df.columns.duplicated()][default_stock_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .sum(axis=1)
            if default_stock_cols else 0
        )
        return out, warehouse_stock_cols

    def load_pricelist_with_warehouses(file):
        xls = pd.ExcelFile(file)
        sheets = [s for s in xls.sheet_names if s.upper() in VALID_PRICELIST_SHEETS]
        frames = []
        merged_map = {}
        for s in sheets:
            part, part_map = parse_pricelist_sheet_with_warehouses(xls, s)
            frames.append(part)
            merged_map.update(part_map)

        if not frames:
            return pd.DataFrame(), {}

        out = pd.concat(frames, ignore_index=True)
        out = out.loc[:, ~out.columns.duplicated()]
        out = out.drop_duplicates(subset=["SKU NO", "KODEBARANG"], keep="first").reset_index(drop=True)
        return out, merged_map

    # =========================================================
    # SALES PIVOT
    # =========================================================

    def load_sales_pivot(file) -> pd.DataFrame:
        raw = pd.read_excel(file, header=1).copy()
        raw = raw.iloc[1:].copy().reset_index(drop=True)
        raw.columns = [str(c).strip().upper() for c in raw.columns]
        raw = raw.loc[:, ~pd.Index(raw.columns).duplicated()].copy()

        team_col = next((c for c in raw.columns if c == "TEAM" or "TEAM" in c), None)
        kode_barang_col = next((c for c in raw.columns if "KODE BARANG" in c or "KODEBARANG" in c), None)
        spesifikasi_col = next((c for c in raw.columns if "SPESIFIKASI" in c), None)
        qty_col = next((c for c in raw.columns if c == "QTY" or "QTY" in c or "PCS" in c or "TERJUAL" in c), None)
        tgl_col = next((c for c in raw.columns if c == "TGL" or "TGL" in c or "DATE" in c), None)
        country_col = next((c for c in raw.columns if c == "COUNTRY" or "COUNTRY" in c), None)
        product_col = next((c for c in raw.columns if c == "PRODUCT" or "PRODUCT" in c), None)
        gp_m0_col = next((c for c in raw.columns if c == "GP M0" or "GPM0" in c or "GP M0" in c.replace("_", " ")), None)
        m0_col = next((c for c in raw.columns if c == "M0" or "M0" in c), None)
        m3_col = next((c for c in raw.columns if c == "M3" or "M3" in c), None)
        harga_akhir_col = next((c for c in raw.columns if "HARGA AKHIR" in c or "HARGAAKHIR" in c), None)

        required = {"TEAM": team_col, "KODE BARANG": kode_barang_col, "QTY": qty_col, "TGL": tgl_col}
        missing = [label for label, col in required.items() if col is None]
        if missing:
            raise ValueError(
                f"Format SALES PIVOT tidak cocok. Kolom wajib tidak ditemukan: {missing}. "
                f"Kolom yang terbaca: {list(raw.columns)}. "
                "Pastikan header ada di row 2 dan data mulai row 4."
            )

        use_cols = [team_col, kode_barang_col, qty_col, tgl_col]
        for extra_col in [country_col, product_col, gp_m0_col, spesifikasi_col, m0_col, m3_col, harga_akhir_col]:
            if extra_col is not None and extra_col not in use_cols:
                use_cols.append(extra_col)

        df = raw[use_cols].copy()
        rename_map = {
            team_col: "TEAM_RAW",
            kode_barang_col: "KODE BARANG",
            qty_col: "QTY",
            tgl_col: "TGL",
        }
        if country_col is not None:
            rename_map[country_col] = "COUNTRY"
        if product_col is not None:
            rename_map[product_col] = "PRODUCT"
        if gp_m0_col is not None:
            rename_map[gp_m0_col] = "GP M0"
        if spesifikasi_col is not None:
            rename_map[spesifikasi_col] = "SPESIFIKASI"
        if m0_col is not None:
            rename_map[m0_col] = "M0"
        if m3_col is not None:
            rename_map[m3_col] = "M3"
        if harga_akhir_col is not None:
            rename_map[harga_akhir_col] = "HARGA AKHIR"
        df = df.rename(columns=rename_map)

        if "COUNTRY" not in df.columns:
            df["COUNTRY"] = np.nan
        if "PRODUCT" not in df.columns:
            df["PRODUCT"] = np.nan
        if "GP M0" not in df.columns:
            df["GP M0"] = np.nan
        if "SPESIFIKASI" not in df.columns:
            df["SPESIFIKASI"] = np.nan
        if "M0" not in df.columns:
            df["M0"] = np.nan
        if "M3" not in df.columns:
            df["M3"] = np.nan
        if "HARGA AKHIR" not in df.columns:
            df["HARGA AKHIR"] = np.nan

        df["TEAM"] = normalize_text(df["TEAM_RAW"])
        df["TEAM_KEY"] = df["TEAM_RAW"].apply(normalize_team_code)
        df["COUNTRY"] = normalize_text(df["COUNTRY"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODE BARANG"] = normalize_text(df["KODE BARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])
        df["QTY"] = to_num(df["QTY"]).fillna(0)
        df["GP M0"] = to_num(df["GP M0"])
        df["M0"] = to_num(df["M0"])
        df["M3"] = to_num(df["M3"])
        df["HARGA AKHIR"] = to_num(df["HARGA AKHIR"])
        df["TGL"] = ensure_datetime(df["TGL"]).dt.normalize()

        df = df[df["TEAM"].notna()].copy()
        df = df[df["TEAM_KEY"].notna()].copy()
        df = df[df["KODE BARANG"].notna()].copy()
        df = df[df["QTY"] > 0].copy()
        df = df[df["TGL"].notna()].copy()

        if df.empty:
            return pd.DataFrame(columns=["TEAM", "TEAM_KEY", "COUNTRY", "PRODUCT", "KODE BARANG", "SPESIFIKASI", "QTY", "TGL", "GP M0", "M0", "M3", "HARGA AKHIR"])

        return df[["TEAM", "TEAM_KEY", "COUNTRY", "PRODUCT", "KODE BARANG", "SPESIFIKASI", "QTY", "TGL", "GP M0", "M0", "M3", "HARGA AKHIR"]].sort_values(
            ["TGL", "TEAM", "KODE BARANG"], ascending=[False, True, True]
        ).reset_index(drop=True)


    def build_sales_pivot_alerts(
        sales_pivot: pd.DataFrame,
        pricelist_wh: pd.DataFrame,
        warehouse_stock_cols: dict,
        period: str,
        selected_products=None,
        selected_brands=None,
        selected_segments=None,
        selected_kode_barang=None,
        selected_teams=None,
        start_date=None,
        end_date=None,
    ) -> pd.DataFrame:
        empty_cols = ["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY"]
        if sales_pivot.empty or pricelist_wh.empty:
            return pd.DataFrame(columns=empty_cols)

        base = sales_pivot.copy()
        base["TGL"] = pd.to_datetime(base["TGL"], errors="coerce").dt.normalize()
        base = base[base["TGL"].notna()].copy()

        default_start_ts, default_end_ts = get_period_date_range(base, period)
        start_ts = pd.to_datetime(start_date, errors="coerce") if start_date is not None else pd.NaT
        end_ts = pd.to_datetime(end_date, errors="coerce") if end_date is not None else pd.NaT

        start_ts = default_start_ts if pd.isna(start_ts) else start_ts.normalize()
        end_ts = default_end_ts if pd.isna(end_ts) else end_ts.normalize()

        if pd.notna(start_ts):
            base = base[base["TGL"] >= start_ts]
        if pd.notna(end_ts):
            base = base[base["TGL"] <= end_ts]

        pl = pricelist_wh.copy()
        if "KODEBARANG" not in pl.columns:
            return pd.DataFrame(columns=empty_cols)

        if selected_products:
            pl = pl[pl["PRODUCT"].isin(selected_products)]
        if selected_brands:
            pl = pl[pl["BRAND"].isin(selected_brands)]
        if selected_segments:
            pl = pl[pl["PRICE_SEGMENT"].isin(selected_segments)]
        if pl.empty:
            return pd.DataFrame(columns=empty_cols)

        allowed_codes = set(normalize_text(pl["KODEBARANG"]).dropna().tolist())
        base["KODE BARANG"] = normalize_text(base["KODE BARANG"])
        base = base[base["KODE BARANG"].isin(allowed_codes)].copy()

        if selected_kode_barang:
            selected_kode_barang_norm = set(normalize_text(pd.Series(selected_kode_barang)).dropna().tolist())
            base = base[base["KODE BARANG"].isin(selected_kode_barang_norm)]
        if selected_teams:
            selected_teams_norm = set(normalize_text(pd.Series(selected_teams)).dropna().tolist())
            base = base[base["TEAM"].isin(selected_teams_norm)]

        if base.empty:
            return pd.DataFrame(columns=empty_cols)

        base = (
            base.groupby(["TEAM", "TEAM_KEY", "KODE BARANG"], as_index=False)
            .agg(SPESIFIKASI=("SPESIFIKASI", "first"), QTY=("QTY", "sum"))
        )

        merged = base.merge(pl, how="left", left_on="KODE BARANG", right_on="KODEBARANG")

        allowed_ready_team_keys = {"JKT1A", "JKT3A", "JKT3B", "JKT3C", "JKT4A", "JKT4B"}

        def get_stock_cols_by_team(team_key):
            if pd.isna(team_key):
                return []
            lookup_key = normalize_team_lookup_key(team_key)

            exact_cols = warehouse_stock_cols.get(lookup_key, [])
            if isinstance(exact_cols, str):
                exact_cols = [exact_cols]
            exact_cols = [c for c in exact_cols if c in merged.columns]
            if exact_cols:
                return exact_cols

            matched = []
            for wh_key, cols in warehouse_stock_cols.items():
                if normalize_team_lookup_key(wh_key) == lookup_key:
                    matched.extend(cols if isinstance(cols, list) else [cols])
            return [c for c in matched if c in merged.columns]

        def sum_stock_from_cols(row, cols):
            if not cols:
                return 0.0
            values = pd.to_numeric(pd.Series([row.get(c, 0) for c in cols]), errors="coerce").fillna(0)
            return float(values.sum())

        team_keys = sorted([
            str(k).strip().upper()
            for k in warehouse_stock_cols.keys()
            if str(k).strip().upper() != "DEFAULT"
        ])

        def get_current_stock(row):
            cols = get_stock_cols_by_team(row.get("TEAM_KEY"))
            return sum_stock_from_cols(row, cols)

        def get_ready_warehouses(row):
            current_team_key = normalize_team_lookup_key(row.get("TEAM_KEY"))
            ready_list = []

            for team_code in team_keys:
                team_key = normalize_team_lookup_key(team_code)
                if team_key == current_team_key:
                    continue
                if team_key not in allowed_ready_team_keys:
                    continue

                cols = warehouse_stock_cols.get(team_code, [])
                cols = cols if isinstance(cols, list) else [cols]
                cols = [c for c in cols if c in merged.columns]
                total_stock = sum_stock_from_cols(row, cols)
                if total_stock > 0:
                    ready_list.append((team_code.replace(" ", ""), int(round(total_stock))))

            ready_list = sorted(ready_list, key=lambda x: x[0])
            return ", ".join([f"{code} ({qty})" for code, qty in ready_list])

        merged["SPESIFIKASI"] = merged["SPESIFIKASI_x"].fillna(merged.get("SPESIFIKASI_y", ""))
        merged["STOK"] = merged.apply(get_current_stock, axis=1)
        merged["GUDANG READY"] = merged.apply(get_ready_warehouses, axis=1)
        merged["KET"] = np.where(
            (to_num(merged["QTY"]).fillna(0) > to_num(merged["STOK"]).fillna(0)) &
            (to_num(merged["STOK"]).fillna(0) <= 0) &
            (merged["GUDANG READY"].astype(str).str.strip() != ""),
            "REFILL",
            np.where(
                (to_num(merged["QTY"]).fillna(0) > to_num(merged["STOK"]).fillna(0)),
                "CEK",
                ""
            )
        )

        merged["KEBUTUHAN_STOK"] = to_num(merged["QTY"]).fillna(0) - to_num(merged["STOK"]).fillna(0)
        merged["PRIORITAS_STOK"] = np.where(
            merged["KET"].astype(str).str.upper().eq("REFILL"),
            2,
            np.where(merged["KET"].astype(str).str.upper().eq("CEK"), 1, 0)
        )

        out = merged[["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY", "KEBUTUHAN_STOK", "PRIORITAS_STOK"]].copy()
        out["QTY"] = pd.to_numeric(out["QTY"], errors="coerce").fillna(0).round(0).astype(int)
        out["STOK"] = pd.to_numeric(out["STOK"], errors="coerce").fillna(0).round(0).astype(int)
        out["KEBUTUHAN_STOK"] = pd.to_numeric(out["KEBUTUHAN_STOK"], errors="coerce").fillna(0)

        out = out.sort_values(
            ["PRIORITAS_STOK", "KEBUTUHAN_STOK", "QTY", "TEAM", "KODE BARANG"],
            ascending=[False, False, False, True, True]
        ).reset_index(drop=True)

        return out[["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY"]]

    def render_sales_pivot_alert_table(df: pd.DataFrame):
        if df.empty:
            st.info("Analisa Sales vs Stok 05 OLR belum menemukan data.")
            return

        show_df = df.copy()
        for col in ["QTY", "STOK"]:
            show_df[col] = pd.to_numeric(show_df[col], errors="coerce").fillna(0).round(0).astype(int)

        html = []
        html.append('<div class="alert-wrap"><table class="alert-table"><thead><tr>')
        for col in show_df.columns:
            html.append(f"<th>{col}</th>")
        html.append("</tr></thead><tbody>")

        for _, row in show_df.iterrows():
            html.append("<tr>")
            for col in show_df.columns:
                cls = ""
                if col in ["STOK", "KET"] and str(row.get("KET", "")).strip().upper() in ["REFILL", "CEK"]:
                    cls = ' class="bg-red"'
                html.append(f"<td{cls}>{row[col]}</td>")
            html.append("</tr>")
        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)


    def build_sku_gp_besar_table(sales_pivot: pd.DataFrame, stock: pd.DataFrame, selected_products=None) -> pd.DataFrame:
        columns = ["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "GP", "STOK"]
        if stock.empty:
            return pd.DataFrame(columns=columns)

        stock_base = stock.copy()
        for col in ["KODEBARANG", "SPESIFIKASI", "PRODUCT", "PRICE", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"]:
            if col not in stock_base.columns:
                stock_base[col] = np.nan if col in ["SPESIFIKASI", "PRODUCT"] else 0

        if selected_products:
            stock_base = stock_base[stock_base["PRODUCT"].isin(selected_products)].copy()

        stock_base["M3_VAL"] = to_num(stock_base.get("PRICE", np.nan)).fillna(0)
        stock_base["M0_VAL"] = 0
        if not sales_pivot.empty and "KODE BARANG" in sales_pivot.columns:
            sales_m0_lookup = sales_pivot.copy()
            sales_m0_lookup["M0_VAL"] = to_num(sales_m0_lookup.get("M0", np.nan)).fillna(0)
            sales_m0_lookup["KODE BARANG"] = normalize_text(sales_m0_lookup["KODE BARANG"])
            sales_m0_lookup = (
                sales_m0_lookup.groupby("KODE BARANG", as_index=False)
                .agg(M0_VAL=("M0_VAL", "max"))
            )
            stock_base = stock_base.merge(
                sales_m0_lookup,
                how="left",
                left_on="KODEBARANG",
                right_on="KODE BARANG",
                suffixes=("", "_sales")
            )
            stock_base["M0_VAL"] = to_num(stock_base.get("M0_VAL_sales", stock_base.get("M0_VAL"))).fillna(
                to_num(stock_base.get("M0_VAL")).fillna(0)
            )
            drop_cols = [c for c in ["KODE BARANG", "M0_VAL_sales"] if c in stock_base.columns]
            if drop_cols:
                stock_base = stock_base.drop(columns=drop_cols)

        stock_base["STOK"] = (
            to_num(stock_base.get("STOK_DIV03", 0)).fillna(0) +
            to_num(stock_base.get("STOK_DIV04", 0)).fillna(0) +
            to_num(stock_base.get("STOK_DIV05", 0)).fillna(0)
        )
        stock_base["GP"] = stock_base["M3_VAL"] - stock_base["M0_VAL"]

        stock_base["KODE BARANG"] = normalize_text(stock_base["KODEBARANG"])
        stock_base["SPESIFIKASI"] = normalize_text(stock_base["SPESIFIKASI"])
        stock_base["PRODUCT"] = normalize_text(stock_base["PRODUCT"])

        out = stock_base[["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3_VAL", "M0_VAL", "GP", "STOK"]].copy()
        out = out.rename(columns={"M3_VAL": "M3", "M0_VAL": "M0"})

        # skip M0 kosong / 0
        out = out[to_num(out["M0"]).fillna(0) > 0].copy()

        out = out[
            (to_num(out["GP"]).fillna(0) > 0) &
            (to_num(out["STOK"]).fillna(0) > 0)
        ].copy()

        out = out.sort_values(["GP", "STOK", "KODE BARANG"], ascending=[False, False, True]).reset_index(drop=True)
        return out[columns]


    def build_sku_top_gp_table(sales_pivot: pd.DataFrame, stock: pd.DataFrame, selected_products=None) -> pd.DataFrame:
        columns = ["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "QTY", "GP M0"]
        if sales_pivot.empty:
            return pd.DataFrame(columns=columns)

        sales_base = sales_pivot.copy()
        for col in ["PRODUCT", "SPESIFIKASI", "KODE BARANG", "M3", "M0", "QTY", "GP M0"]:
            if col not in sales_base.columns:
                sales_base[col] = np.nan

        sales_base["PRODUCT"] = normalize_text(sales_base["PRODUCT"])
        sales_base["KODE BARANG"] = normalize_text(sales_base["KODE BARANG"])
        sales_base["SPESIFIKASI"] = normalize_text(sales_base["SPESIFIKASI"])
        sales_base["M3_VAL"] = to_num(sales_base.get("M3", np.nan)).fillna(0)
        sales_base["M0_VAL"] = to_num(sales_base.get("M0", np.nan)).fillna(0)
        sales_base["GP_M0_VAL"] = to_num(sales_base.get("GP M0", np.nan)).fillna(0)
        sales_base["QTY"] = to_num(sales_base.get("QTY", 0)).fillna(0)

        if selected_products:
            sales_base = sales_base[sales_base["PRODUCT"].isin(selected_products)].copy()

        sales_base = sales_base[(sales_base["KODE BARANG"].notna()) & (sales_base["PRODUCT"].notna())].copy()
        sales_base = sales_base[sales_base["GP_M0_VAL"] > 0].copy()

        if sales_base.empty:
            return pd.DataFrame(columns=columns)

        # Samakan dengan Pivot Excel:
        # - filter PRODUCT dari file sales
        # - row label berdasarkan KODE BARANG
        # - value adalah SUM dari kolom GP M0 asli di file sales
        # Kolom lain diambil representatif per KODE BARANG agar tampilan card tetap informatif.
        meta = (
            sales_base.sort_values(["KODE BARANG", "QTY", "GP_M0_VAL"], ascending=[True, False, False])
            .drop_duplicates(subset=["KODE BARANG"], keep="first")
            [["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3_VAL", "M0_VAL"]]
            .copy()
        )

        pivot_like = (
            sales_base.groupby(["KODE BARANG"], dropna=False, as_index=False)
            .agg(QTY=("QTY", "sum"), **{"GP M0": ("GP_M0_VAL", "sum")})
        )

        out = pivot_like.merge(meta, how="left", on="KODE BARANG")
        out = out.rename(columns={"M3_VAL": "M3", "M0_VAL": "M0"})
        out = out[["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "QTY", "GP M0"]]
        out = out.sort_values(["GP M0", "QTY", "KODE BARANG"], ascending=[False, False, True]).reset_index(drop=True)
        return out[columns]


    def render_simple_card_table(df: pd.DataFrame, title: str):
        st.markdown(f"### {title}")
        if df.empty:
            st.info(f"{title} belum menemukan data.")
            return

        show_df = df.copy()
        for col in ["M3", "M0", "GP", "GP TOTAL", "GP M0"]:
            if col in show_df.columns:
                show_df[col] = show_df[col].apply(format_thousands_id)
        for col in ["QTY", "STOK"]:
            if col in show_df.columns:
                show_df[col] = show_df[col].apply(format_units_id)

        html = []
        html.append('<div class="main-fixed-wrap"><table class="main-fixed"><thead><tr>')
        for col in show_df.columns:
            width_style = ""
            if col in ["M3", "M0", "GP", "QTY", "STOK"]:
                width_style = ' style="min-width:70px;max-width:70px;"'
            html.append(f"<th{width_style}>{col}</th>")
        html.append("</tr></thead><tbody>")

        for _, row in show_df.iterrows():
            html.append("<tr>")
            for col in show_df.columns:
                width_style = ""
                if col in ["M3", "M0", "GP", "QTY", "STOK"]:
                    width_style = ' style="min-width:70px;max-width:70px;"'
                html.append(f"<td{width_style}>{'' if pd.isna(row[col]) else row[col]}</td>")
            html.append("</tr>")

        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)

    # =========================================================
    # BUILD TABLES
    # =========================================================
    def build_master(sales: pd.DataFrame, stock: pd.DataFrame) -> pd.DataFrame:
        df = sales.merge(stock, how="left", on="MERGE_KEY", suffixes=("_sales", "_stock"))

        for col in ["PRICE", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05", "CATEGORY", "PRICE_SEGMENT"]:
            if col not in df.columns:
                df[col] = np.nan

        df["KODEBARANG"] = normalize_text(df["KODE BARANG"])
        df["SPESIFIKASI_FINAL"] = df["SPESIFIKASI_sales"].fillna(df.get("SPESIFIKASI_stock"))
        df["PRODUCT_FINAL"] = df["PRODUCT_sales"].fillna(df.get("PRODUCT_stock"))
        df["BRAND"] = normalize_text(df["BRAND"])
        df["QTY"] = to_num(df["QTY"]).fillna(0)

        stock_map = {"DIV03": "STOK_DIV03", "DIV04": "STOK_DIV04", "DIV05": "STOK_DIV05"}
        df["STOK_DIVISI"] = df.apply(lambda r: pd.to_numeric(r.get(stock_map[r["DIVISION"]]), errors="coerce"), axis=1).fillna(0)
        return df

    def build_segment_table(df, period, comparison_division_label="03 OLP"):
        tmp = df[df["PERIOD"] == period].copy()
        tmp["SEGMENT"] = tmp["PRICE"].apply(price_segment)
        seg = tmp.groupby(["SEGMENT", "DIVISION"])["QTY"].sum().unstack().fillna(0).reset_index()
        for div in DIVISIONS:
            if div not in seg.columns:
                seg[div] = 0
        seg = seg[["SEGMENT", "DIV03", "DIV04", "DIV05"]].copy()
        seg = seg.sort_values("SEGMENT", key=lambda s: s.map(segment_sort_key)).reset_index(drop=True)
        seg.columns = ["SEGMENT", "03 OLP", "04 MOD", "05 OLR"]

        compare_col = comparison_division_label if comparison_division_label in ["03 OLP", "04 MOD"] else "03 OLP"
        seg["DELTA"] = to_num(seg["05 OLR"]).fillna(0) - to_num(seg[compare_col]).fillna(0)
        return seg

    def build_brand_table(df, period, comparison_division_label="03 OLP"):
        brand = df[df["PERIOD"] == period].copy()
        brand = brand.groupby(["BRAND", "DIVISION"])["QTY"].sum().unstack().fillna(0).reset_index()
        for div in DIVISIONS:
            if div not in brand.columns:
                brand[div] = 0
        brand = brand[["BRAND", "DIV03", "DIV04", "DIV05"]].copy()
        brand["TOTAL"] = brand[["DIV03", "DIV04", "DIV05"]].sum(axis=1)
        brand = brand.sort_values(["TOTAL", "BRAND"], ascending=[False, True]).drop(columns=["TOTAL"])
        brand.columns = ["BRAND", "03 OLP", "04 MOD", "05 OLR"]

        compare_col = comparison_division_label if comparison_division_label in ["03 OLP", "04 MOD"] else "03 OLP"
        brand["DELTA"] = to_num(brand["05 OLR"]).fillna(0) - to_num(brand[compare_col]).fillna(0)
        return brand

    def render_left_table(df, title, selected_division="05 OLR", use_card=True):
        def is_number(v):
            return isinstance(v, (int, float, np.integer, np.floating)) and not pd.isna(v)

        def fmt_number(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        html = []
        if use_card:
            html.append("""
            <div class="table-card">
              <div style="font-weight:700;font-size:16px;margin-bottom:8px;">""" + title + """</div>
              <div style="overflow-x:auto;">
                <table style="border-collapse:collapse;width:100%;font-size:12px;">
            """)
        else:
            html.append("""
              <div style="font-weight:700;font-size:16px;margin-bottom:8px;">""" + title + """</div>
              <div style="overflow-x:auto;">
                <table style="border-collapse:collapse;width:100%;font-size:12px;">
            """)
        html.append("<thead><tr>")
        for col in df.columns:
            html.append(f'<th style="border:1px solid #2b2b2b;background:#f3f4f6;padding:6px;text-align:left;">{col}</th>')
        html.append("</tr></thead><tbody>")

        compare_cols = ["03 OLP", "04 MOD", "05 OLR"]
        has_compare_cols = all(c in df.columns for c in compare_cols)

        for _, row in df.iterrows():
            html.append("<tr>")

            losing_selected = False
            if has_compare_cols and selected_division in compare_cols:
                current_val = row[selected_division]
                other_vals = [row[c] for c in compare_cols if c != selected_division]
                if is_number(current_val):
                    numeric_others = [v for v in other_vals if is_number(v)]
                    if numeric_others:
                        losing_selected = any(float(current_val) < float(v) for v in numeric_others)

            for col in df.columns:
                val = row[col]
                try:
                    if pd.notna(val) and isinstance(val, (int, float, np.integer, np.floating)):
                        display = fmt_number(val)
                    else:
                        display = "" if pd.isna(val) else str(val)
                except Exception:
                    display = str(val)

                style = 'border:1px solid #2b2b2b;padding:6px;text-align:left;'
                if col == selected_division and losing_selected:
                    style += 'color:#c62828;font-weight:700;background:#ffebee;'
                if col == "DELTA" and is_number(val) and float(val) < 0:
                    style += 'color:#c62828;font-weight:700;background:#ffebee;'

                html.append(f'<td style="{style}">{display}</td>')
            html.append("</tr>")

        if use_card:
            html.append("</tbody></table></div></div>")
        else:
            html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)

    def build_main_table_filtered(
        df: pd.DataFrame,
        period: str,
        comparison_division_label: str,
        selected_segments=None,
        selected_brands=None,
        selected_products=None,
    ) -> pd.DataFrame:
        base = df[df["PERIOD"] == period].copy()

        if selected_segments:
            base = base[base["PRICE"].apply(price_segment).isin(selected_segments)]
        if selected_brands:
            base = base[base["BRAND"].isin(selected_brands)]
        if selected_products:
            base = base[base["PRODUCT_FINAL"].isin(selected_products)]

        qty = (
            base.groupby(["KODEBARANG", "SPESIFIKASI_FINAL", "PRICE", "DIVISION"], as_index=False)["QTY"]
            .sum()
            .pivot(index=["KODEBARANG", "SPESIFIKASI_FINAL", "PRICE"], columns="DIVISION", values="QTY")
            .fillna(0)
            .reset_index()
        )

        for div in DIVISIONS:
            if div not in qty.columns:
                qty[div] = 0

        product_brand = (
            base[["KODEBARANG", "PRODUCT_FINAL", "BRAND"]]
            .dropna(subset=["KODEBARANG"])
            .drop_duplicates(subset=["KODEBARANG"], keep="first")
            .copy()
        )

        stock_label_map = {"03 OLP": "STOK_DIV03", "04 MOD": "STOK_DIV04", "05 OLR": "STOK_DIV05"}
        stok_col = "STOK_DIV05"

        stock_df = (
            base[["KODEBARANG", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"]]
            .dropna(subset=["KODEBARANG"])
            .drop_duplicates(subset=["KODEBARANG"], keep="first")
            .copy()
        )
        stock_df["STOK"] = to_num(stock_df[stok_col]).fillna(0)

        out = qty.merge(product_brand, how="left", on="KODEBARANG")
        out = out.merge(stock_df, how="left", on="KODEBARANG")
        out["STOK"] = to_num(out["STOK"]).fillna(0)
        out["PRICE"] = to_num(out["PRICE"]).fillna(0)
        out["STOK_DIV03"] = to_num(out["STOK_DIV03"]).fillna(0)
        out["STOK_DIV04"] = to_num(out["STOK_DIV04"]).fillna(0)
        out["STOK_DIV05"] = to_num(out["STOK_DIV05"]).fillna(0)

        out = out.rename(columns={
            "PRODUCT_FINAL": "PRODUCT",
            "SPESIFIKASI_FINAL": "SPESIFIKASI",
            "PRICE": "M3",
            "DIV03": "03 OLP",
            "DIV04": "04 MOD",
            "DIV05": "05 OLR",
        })

        ordered_cols = [
            "KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI", "M3",
            "03 OLP", "04 MOD", "05 OLR", "STOK",
            "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"
        ]
        for col in ordered_cols:
            if col not in out.columns:
                out[col] = 0 if col not in ["KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI"] else ""

        out["M3"] = (to_num(out["M3"]).fillna(0) / 1000).round(0)
        compare_map = {"03 OLP": "03 OLP", "04 MOD": "04 MOD"}
        compare_col = compare_map.get(comparison_division_label, "03 OLP")
        out["DELTA"] = to_num(out["05 OLR"]).fillna(0) - to_num(out[compare_col]).fillna(0)
        return out[ordered_cols + ["DELTA"]].sort_values(["DELTA", compare_col, "05 OLR"], ascending=[True, False, True]).reset_index(drop=True)

    def render_main_table_dynamic(df: pd.DataFrame, comparison_division_label: str):
        display_df = df.copy()

        compare_cols = ["03 OLP", "04 MOD", "05 OLR"]
        stock_hidden_map = {"03 OLP": "STOK_DIV03", "04 MOD": "STOK_DIV04", "05 OLR": "STOK_DIV05"}
        selected_stock_hidden = "STOK_DIV05"

        def losing_division(row):
            current_val = row.get("05 OLR", 0)
            compare_val = row.get(comparison_division_label, 0)
            try:
                return float(current_val) < float(compare_val)
            except Exception:
                return False

        def stok_problem(row):
            try:
                stok_selected = float(row.get("STOK", 0))
                qty_selected = float(row.get("05 OLR", 0))
                other_stock_cols = [stock_hidden_map[c] for c in compare_cols if c != "05 OLR"]
                other_stock_values = [float(row.get(c, 0)) for c in other_stock_cols]
                cond_a = stok_selected < qty_selected
                cond_b = stok_selected == 0 and qty_selected > 0 and any(v > 0 for v in other_stock_values)
                return cond_a or cond_b
            except Exception:
                return False

        display_df["_LOSS_DIVISION_FLAG"] = display_df.apply(losing_division, axis=1)
        display_df["_STOK_ALERT_FLAG"] = display_df.apply(stok_problem, axis=1)

        visible_df = display_df[["KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI", "M3", "03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]].copy()
        visible_df["M3"] = pd.to_numeric(visible_df["M3"], errors="coerce").fillna(0).round(0).astype(int)
        for col in ["03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]:
            numeric_col = pd.to_numeric(visible_df[col], errors="coerce").fillna(0).round(0)
            visible_df[col] = numeric_col.astype(int)

        def fmt_value(val, col_name):
            if col_name in ["M3", "03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]:
                try:
                    return f"{int(val)}"
                except Exception:
                    return "0"
            return "" if pd.isna(val) else str(val)

        html = []
        html.append('<div class="main-fixed-wrap"><table class="main-fixed"><thead><tr>')
        for col in visible_df.columns:
            html.append(f"<th>{col}</th>")
        html.append("</tr></thead><tbody>")

        for idx, row in visible_df.iterrows():
            html.append("<tr>")
            original = display_df.loc[idx]
            for col in visible_df.columns:
                cls = ""
                if col == "05 OLR" and bool(original["_LOSS_DIVISION_FLAG"]):
                    cls = ' class="bg-red"'
                if col == "DELTA":
                    try:
                        if float(original.get("DELTA", 0)) < 0:
                            cls = ' class="bg-red"'
                    except Exception:
                        pass
                if col == "STOK" and bool(original["_STOK_ALERT_FLAG"]):
                    cls = ' class="bg-red"'
                html.append(f"<td{cls}>{fmt_value(row[col], col)}</td>")
            html.append("</tr>")

        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)
        return visible_df

    # =========================================================
    # UI
    # =========================================================
    st.title("Analisa Produk & Stok")

    st.markdown('<div class="upload-card-wrap">', unsafe_allow_html=True)
    st.markdown("### Upload File")

    upload_col1, upload_col2, upload_col3 = st.columns(3)
    with upload_col1:
        st.markdown("**Upload MPLSSR**")
        mplssr_file = st.file_uploader("Upload MPLSSR", type=["xlsx", "xls"], key="upload_mplssr_main", label_visibility="collapsed")
        st.caption("200MB per file • XLSX, XLS")

    with upload_col2:
        st.markdown("**Upload Pricelist**")
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx", "xls"], key="upload_pricelist_main", label_visibility="collapsed")
        st.caption("200MB per file • XLSX, XLS")

    with upload_col3:
        st.markdown("**Upload Penjualan**")
        sales_pivot_file = st.file_uploader("Upload Penjualan", type=["xlsx", "xls"], key="upload_sales_pivot_main", label_visibility="collapsed")
        st.caption("200MB per file • XLSX, XLS")

    all_required_uploaded = all([mplssr_file is not None, pricelist_file is not None, sales_pivot_file is not None])

    if not all_required_uploaded:
        st.info("Silakan upload MPLSSR, Pricelist, dan Penjualan.")
    st.markdown('</div>', unsafe_allow_html=True)

    if all_required_uploaded:
        current_upload_signature = (
            getattr(mplssr_file, "name", None),
            getattr(pricelist_file, "name", None),
            getattr(sales_pivot_file, "name", None),
        )

        if st.session_state.get("processed_upload_signature") != current_upload_signature:
            try:
                sales = load_mplssr(mplssr_file)
                stock = load_pricelist(pricelist_file)
                master = build_master(sales, stock)
                pricelist_wh, warehouse_stock_cols = load_pricelist_with_warehouses(pricelist_file)
                sales_pivot = load_sales_pivot(sales_pivot_file)
                st.session_state["processed_data"] = {
                    "sales": sales,
                    "stock": stock,
                    "master": master,
                    "pricelist_wh": pricelist_wh,
                    "warehouse_stock_cols": warehouse_stock_cols,
                    "sales_pivot": sales_pivot,
                }
                st.session_state["processed_upload_signature"] = current_upload_signature
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
                st.stop()

    if "processed_data" not in st.session_state:
        st.stop()

    sales = st.session_state["processed_data"]["sales"]
    stock = st.session_state["processed_data"]["stock"]
    master = st.session_state["processed_data"]["master"]
    pricelist_wh = st.session_state["processed_data"]["pricelist_wh"]
    warehouse_stock_cols = st.session_state["processed_data"]["warehouse_stock_cols"]
    sales_pivot = st.session_state["processed_data"]["sales_pivot"]

    product_options = sorted(master["PRODUCT_FINAL"].dropna().unique().tolist())
    default_product = ["LAPTOP R"] if "LAPTOP R" in product_options else []

    st.markdown('<div class="upload-card-wrap">', unsafe_allow_html=True)
    st.markdown("### Filter Dashboard")
    with st.form("unified_filter_form"):
        fcol1, fcol2, fcol3, fcol4, fcol5, fcol6 = st.columns([1.2, 1.2, 1, 1.2, 1.1, 0.6])
        with fcol1:
            selected_products = st.multiselect("Product", product_options, default=default_product)
        with fcol2:
            selected_brands = st.multiselect("Brand", sorted(master["BRAND"].dropna().unique().tolist()))
        with fcol3:
            selected_period = st.selectbox("Period", PERIODS, index=0)
        with fcol4:
            selected_segments = st.multiselect("Range Harga", [s[2] for s in PRICE_SEGMENTS] + ["UNKNOWN"])
        with fcol5:
            comparison_division = st.selectbox("Perbandingan", ["03 OLP", "04 MOD", "05 OLR"], index=0)
        with fcol6:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            apply_filter = st.form_submit_button("PROSES")
    st.markdown('</div>', unsafe_allow_html=True)

    filtered = master.copy()
    if selected_products:
        filtered = filtered[filtered["PRODUCT_FINAL"].isin(selected_products)]
    if selected_brands:
        filtered = filtered[filtered["BRAND"].isin(selected_brands)]
    if selected_segments:
        filtered = filtered[filtered["PRICE"].apply(price_segment).isin(selected_segments)]

    if filtered.empty:
        st.warning("Data kosong setelah filter diterapkan.")
        st.stop()

    with st.container(border=True):
        st.markdown("### Analisa Segment vs Divisi Lain")
        left, right = st.columns(2)
        with left:
            render_left_table(build_segment_table(filtered, selected_period, comparison_division), f"Segmentasi Harga - {selected_period}", selected_division=comparison_division, use_card=False)
        with right:
            render_left_table(build_brand_table(filtered, selected_period, comparison_division), f"Segmentasi Brand - {selected_period}", selected_division=comparison_division, use_card=False)

    with st.container(border=True):
        st.markdown("### Analisa SKU vs Divisi Lain")
        main_table_export = build_main_table_filtered(
            filtered,
            selected_period,
            comparison_division,
            selected_segments=selected_segments,
            selected_brands=selected_brands,
            selected_products=selected_products,
        )
        main_table_export = render_main_table_dynamic(main_table_export, comparison_division)

    stok_kode_barang_options = sorted(sales_pivot["KODE BARANG"].dropna().unique().tolist()) if not sales_pivot.empty and "KODE BARANG" in sales_pivot.columns else []
    stok_team_options = sorted(normalize_text(sales_pivot["TEAM"]).dropna().unique().tolist()) if not sales_pivot.empty and "TEAM" in sales_pivot.columns else []
    stok_min_date = sales_pivot["TGL"].min() if not sales_pivot.empty and "TGL" in sales_pivot.columns else None
    stok_max_date = sales_pivot["TGL"].max() if not sales_pivot.empty and "TGL" in sales_pivot.columns else None
    stok_min_date_value = pd.to_datetime(stok_min_date, errors="coerce").date() if pd.notna(stok_min_date) else None
    stok_max_date_value = pd.to_datetime(stok_max_date, errors="coerce").date() if pd.notna(stok_max_date) else None

    if "stok_products" not in st.session_state:
        st.session_state["stok_products"] = selected_products
    if "stok_period" not in st.session_state:
        st.session_state["stok_period"] = selected_period if selected_period in PERIODS else PERIODS[0]
    if "stok_kode_barang" not in st.session_state:
        st.session_state["stok_kode_barang"] = []
    if "stok_teams" not in st.session_state:
        st.session_state["stok_teams"] = []

    period_default_start, period_default_end = get_period_date_range(sales_pivot, st.session_state["stok_period"])
    period_default_start_value = pd.to_datetime(period_default_start, errors="coerce").date() if pd.notna(period_default_start) else stok_min_date_value
    period_default_end_value = pd.to_datetime(period_default_end, errors="coerce").date() if pd.notna(period_default_end) else stok_max_date_value

    current_start = pd.to_datetime(st.session_state.get("stok_start_date"), errors="coerce")
    current_end = pd.to_datetime(st.session_state.get("stok_end_date"), errors="coerce")

    if pd.isna(current_start):
        st.session_state["stok_start_date"] = period_default_start_value
    else:
        current_start_value = current_start.date()
        if stok_min_date_value is not None and (current_start_value < stok_min_date_value or current_start_value > stok_max_date_value):
            st.session_state["stok_start_date"] = period_default_start_value

    if pd.isna(current_end):
        st.session_state["stok_end_date"] = period_default_end_value
    else:
        current_end_value = current_end.date()
        if stok_min_date_value is not None and (current_end_value < stok_min_date_value or current_end_value > stok_max_date_value):
            st.session_state["stok_end_date"] = period_default_end_value

    if (
        st.session_state.get("stok_start_date") is not None
        and st.session_state.get("stok_end_date") is not None
        and st.session_state["stok_start_date"] > st.session_state["stok_end_date"]
    ):
        st.session_state["stok_start_date"] = period_default_start_value
        st.session_state["stok_end_date"] = period_default_end_value

    with st.container(border=True):
        st.markdown("### Analisa Sales vs Stok 05 OLR")
        with st.form("stok_filter_form"):
            col1, col2 = st.columns(2)
            with col1:
                stok_products = st.multiselect(
                    "Product",
                    sorted(pricelist_wh["PRODUCT"].dropna().unique()),
                    default=st.session_state.get("stok_products", selected_products),
                )
            with col2:
                current_stok_period = st.session_state.get("stok_period", selected_period if selected_period in PERIODS else PERIODS[0])
                stok_period = st.selectbox(
                    "Period",
                    PERIODS,
                    index=PERIODS.index(current_stok_period) if current_stok_period in PERIODS else 0,
                )

            col3, col4 = st.columns(2)
            with col3:
                stok_kode_barang = st.multiselect(
                    "Kode Barang",
                    stok_kode_barang_options,
                    default=st.session_state.get("stok_kode_barang", []),
                )
            with col4:
                stok_teams = st.multiselect(
                    "Team",
                    stok_team_options,
                    default=st.session_state.get("stok_teams", []),
                )

            col5 = st.columns(1)[0]
            with col5:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                process_stok = st.form_submit_button("PROSES")

        if process_stok:
            new_period_start, new_period_end = get_period_date_range(sales_pivot, stok_period)
            new_period_start_value = pd.to_datetime(new_period_start, errors="coerce").date() if pd.notna(new_period_start) else stok_min_date_value
            new_period_end_value = pd.to_datetime(new_period_end, errors="coerce").date() if pd.notna(new_period_end) else stok_max_date_value

            st.session_state["stok_products"] = stok_products
            st.session_state["stok_period"] = stok_period
            st.session_state["stok_kode_barang"] = stok_kode_barang
            st.session_state["stok_teams"] = stok_teams

            start_value = pd.to_datetime(st.session_state.get("stok_start_date"), errors="coerce")
            end_value = pd.to_datetime(st.session_state.get("stok_end_date"), errors="coerce")
            start_value = start_value.date() if pd.notna(start_value) else new_period_start_value
            end_value = end_value.date() if pd.notna(end_value) else new_period_end_value

            if stok_min_date_value is not None and (start_value < stok_min_date_value or start_value > stok_max_date_value):
                start_value = new_period_start_value
            if stok_min_date_value is not None and (end_value < stok_min_date_value or end_value > stok_max_date_value):
                end_value = new_period_end_value
            if start_value > end_value:
                start_value, end_value = new_period_start_value, new_period_end_value

            st.session_state["stok_start_date"] = start_value
            st.session_state["stok_end_date"] = end_value

        sales_pivot_alerts = build_sales_pivot_alerts(
            sales_pivot,
            pricelist_wh,
            warehouse_stock_cols,
            period=st.session_state["stok_period"],
            selected_products=st.session_state["stok_products"],
            selected_kode_barang=st.session_state.get("stok_kode_barang", []),
            selected_teams=st.session_state.get("stok_teams", []),
        
        )

        render_sales_pivot_alert_table(sales_pivot_alerts)

    if "gp_products" not in st.session_state:
        st.session_state["gp_products"] = default_product.copy()
    top_gp_product_options = sorted(sales_pivot["PRODUCT"].dropna().unique().tolist()) if not sales_pivot.empty and "PRODUCT" in sales_pivot.columns else []
    default_top_gp_product = ["LAPTOP R"] if "LAPTOP R" in top_gp_product_options else []

    if "top_gp_products" not in st.session_state:
        st.session_state["top_gp_products"] = default_top_gp_product.copy()
    if "gp_cards_applied" not in st.session_state:
        st.session_state["gp_cards_applied"] = True

    card_col1, card_col2 = st.columns(2)

    with card_col1:
        with st.container(border=True):
            with st.form("gp_besar_form"):
                gp_filter_col, gp_button_col = st.columns([4, 1])
                with gp_filter_col:
                    gp_product_filter = st.multiselect(
                        "Filter Product - SKU Dengan GP Besar",
                        product_options,
                        default=st.session_state.get("gp_products", []),
                        key="gp_products_filter",
                    )
                with gp_button_col:
                    st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                    process_gp_besar = st.form_submit_button("PROSES")

            if process_gp_besar:
                st.session_state["gp_products"] = gp_product_filter
                st.session_state["gp_cards_applied"] = True

            if st.session_state.get("gp_cards_applied"):
                render_simple_card_table(
                    build_sku_gp_besar_table(
                        sales_pivot=sales_pivot,
                        stock=stock,
                        selected_products=st.session_state.get("gp_products", []),
                    ),
                    "SKU Dengan GP Besar"
                )
            else:
                st.info("Pilih filter product lalu klik PROSES.")

    with card_col2:
        with st.container(border=True):
            with st.form("top_gp_form"):
                top_filter_col, top_button_col = st.columns([4, 1])
                with top_filter_col:
                    top_gp_product_filter = st.multiselect(
                        "Filter Product - SKU Top GP",
                        top_gp_product_options,
                        default=st.session_state.get("top_gp_products", default_top_gp_product),
                        key="top_gp_products_filter",
                    )
                with top_button_col:
                    st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                    process_top_gp = st.form_submit_button("PROSES")

            if process_top_gp:
                st.session_state["top_gp_products"] = top_gp_product_filter
                st.session_state["gp_cards_applied"] = True

            if st.session_state.get("gp_cards_applied"):
                render_simple_card_table(
                    build_sku_top_gp_table(
                        sales_pivot=sales_pivot,
                        stock=stock,
                        selected_products=st.session_state.get("top_gp_products", []),
                    ),
                    "SKU Top GP"
                )
            else:
                st.info("Pilih filter product lalu klik PROSES.")

    st.markdown("<div style='height:120px;'></div>", unsafe_allow_html=True)


# ============================================================
# SIDEBAR ROUTER
# ============================================================
def build_menu() -> str:
    st.sidebar.title(APP_TITLE)

    group = st.sidebar.radio(
        "Menu Utama",
        ["Dashboard", "Update Stok", "Update Harga Normal", "Update Harga Coret", "Submit Campaign", "Analisa Penjualan", "Analisa Produk & Stok"],
        key="sidebar_main_menu",
    )

    if group == "Dashboard":
        route = "dashboard"

    elif group == "Update Stok":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "Bigseller", "Blibli", "Akulaku"],
            key="sidebar_update_stok_menu",
        )
        if child.startswith("Shopee"):
            route = "update_stok_shopee"
        elif child == "TikTokShop":
            route = "update_stok_tiktokshop"
        elif child == "Bigseller":
            route = "update_stok_bigseller"
        elif child == "Blibli":
            route = "update_stok_blibli"
        else:
            route = "update_stok_akulaku"

    elif group == "Update Harga Normal":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "PowerMerchant", "Bigseller", "Blibli", "Akulaku"],
            key="sidebar_harga_normal_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_normal_shopee"
        elif child == "TikTokShop":
            route = "harga_normal_tiktokshop"
        elif child == "PowerMerchant":
            route = "harga_normal_powermerchant"
        elif child == "Bigseller":
            route = "harga_normal_bigseller"
        elif child == "Blibli":
            route = "harga_normal_blibli"
        else:
            route = "harga_normal_akulaku"

    elif group == "Update Harga Coret":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "PowerMerchant"],
            key="sidebar_harga_coret_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_coret_shopee"
        elif child == "TikTokShop":
            route = "harga_coret_tiktokshop"
        else:
            route = "harga_coret_powermerchant"

    elif group == "Submit Campaign":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee", "TikTokShop"],
            key="sidebar_submit_campaign_menu",
        )
        if child == "Shopee":
            route = "submit_campaign_shopee"
        else:
            route = "submit_campaign_tiktokshop"

    elif group == "Analisa Penjualan":
        route = "analisa_penjualan"

    elif group == "Analisa Produk & Stok":
        route = "analisa_produk_stok"

    else:
        route = "dashboard"

    st.sidebar.markdown("---")
    st.sidebar.markdown("<br><br><br>", unsafe_allow_html=True)
    st.sidebar.link_button(
        "Download File Addon",
        "https://drive.google.com/drive/u/0/folders/1r3qVqmm1ALfLGaLuvagAf5EQuMVT0iWI",
        use_container_width=True,
    )

    return route


def main():
    route = build_menu()
    if route == "dashboard":
        render_dashboard()
    elif route == "update_stok_shopee":
        render_update_stok_shopee()
    elif route == "update_stok_tiktokshop":
        render_update_stok_tiktokshop()
    elif route == "update_stok_bigseller":
        render_update_stok_bigseller()
    elif route == "update_stok_blibli":
        render_update_stok_blibli()
    elif route == "update_stok_akulaku":
        render_update_stok_akulaku()
    elif route == "harga_normal_shopee":
        render_harga_normal_shopee()
    elif route == "harga_normal_tiktokshop":
        render_harga_normal_tiktokshop()
    elif route == "harga_normal_powermerchant":
        render_harga_normal_powemerchant()
    elif route == "harga_normal_bigseller":
        render_harga_normal_bigseller()
    elif route == "harga_normal_blibli":
        render_harga_normal_blibli()
    elif route == "harga_normal_akulaku":
        render_harga_normal_akulaku()
    elif route == "harga_coret_shopee":
        render_harga_coret_shopee()
    elif route == "harga_coret_tiktokshop":
        render_harga_coret_tiktokshop()
    elif route == "harga_coret_powermerchant":
        render_harga_coret_powemerchant()
    elif route == "submit_campaign_shopee":
        render_submit_campaign_shopee()
    elif route == "submit_campaign_tiktokshop":
        render_submit_campaign_tiktokshop()
    elif route == "analisa_penjualan":
        render_analisa_penjualan()
    elif route == "analisa_produk_stok":
        render_analisa_produk_stok()
    else:
        st.error("Menu tidak dikenal.")


if __name__ == "__main__":
    main()
